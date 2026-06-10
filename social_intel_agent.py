"""
social_intel_agent.py
---------------------
Social intelligence agent for Smoke Shoppe.

Maintains two caches:
  data/Cigar_Social.xlsx  — per-SKU reputation: pro review scores + Reddit/YouTube sentiment
  data/Cigar_Buzz.xlsx    — new & upcoming cigars generating online buzz

Web search is handled by Anthropic's native web_search_20250305 tool (same as
cigar_researcher.py — no extra API key, billed at $10/1,000 searches).

Optional enrichment (both degrade gracefully if not configured):
  Reddit community data  — set REDDIT_CLIENT_ID + REDDIT_CLIENT_SECRET
  YouTube video data     — set YOUTUBE_API_KEY

Usage:
  # One cigar reputation lookup (prints JSON)
  python social_intel_agent.py "Perdomo BBA Mad. Churchill" "Perdomo"

  # Batch-research all uncached inventory cigars
  python social_intel_agent.py --batch [--limit N] [--since "last 6 months"]

  # Refresh the buzz feed (new/upcoming cigars)
  python social_intel_agent.py --buzz

  # Check cache coverage
  python social_intel_agent.py --status
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from tools.reddit_tool import (
    search_cigars as reddit_search,
    format_for_prompt as reddit_format,
)
from tools.youtube_tool import (
    search_videos as youtube_search,
    format_for_prompt as youtube_format,
)

# Re-use battle-tested helpers from cigar_researcher
from cigar_researcher import _create_with_backoff, _apply_sales_filter

# ── paths ─────────────────────────────────────────────────────────────────────
DATA_DIR       = Path(__file__).parent / "data"
SOCIAL_FILE    = DATA_DIR / "Cigar_Social.xlsx"
BUZZ_FILE      = DATA_DIR / "Cigar_Buzz.xlsx"
INVENTORY_FILE = DATA_DIR / "Smoke_Shoppe_Inventory_Verified.xlsx"
RESEARCH_FILE  = DATA_DIR / "Cigar_Research.xlsx"   # for MSRP cross-reference

# ── XLSX schemas ──────────────────────────────────────────────────────────────
SOCIAL_COLUMNS = [
    "Item Number", "Description", "Brand", "Parent Company",
    "Overall Score", "Quality Score", "Value Score", "Community Score",
    "Review Count", "Reddit Mentions", "YouTube Videos",
    "Top Quotes",
    "Source Breakdown",
    "Source URLs",
    "Research Date", "Notes",
]

BUZZ_COLUMNS = [
    "Name", "Brand", "Parent Company",
    "Announced Date", "Release Status",
    "Buzz Score", "Fit Score",
    "Reddit Mentions", "YouTube Videos",
    "Sentiment", "Summary", "Fit Notes",
    "Source URLs",
    "Research Date",
]

# ── Store fit profile ─────────────────────────────────────────────────────────
# Derived from 18 months of sales data (Nov 2024 – May 2026).
# Methodology: balanced average of quantity-share and revenue-share for each dimension,
# so cheap high-volume items (e.g. Villiger Braniff cigarillos) don't overwhelm
# the picture. Premium tier defined as unit price >= $8.
# Pass this (or a custom string) to refresh_buzz_feed() as fit_profile.
DEFAULT_FIT_PROFILE = """
STORE FIT PROFILE — Smoke Shoppe (Nov 2024–May 2026, balanced qty+revenue weighting)

PRICE — sweet spot for new premium additions:
  $8–12    40% balanced share  (high volume, moderate revenue)
  $12–18   20% balanced share  ← ideal zone for a new premium addition
  $18–25    6% balanced share  (small but real; acceptable for standout cigars)
  $25+      2% balanced share  — very hard to move; avoid unless extraordinary buzz
  Revenue-weighted avg for premium ($8+) purchases: $12.15
  New cigars above $20 require strong justification; above $30 are nearly impossible.

WRAPPER (balanced qty+revenue):
  #1 Maduro / Capa Negra     45.5%  ← dominant; dark wrappers far outperform light
  #2 Sumatra / Natural       17.2%
  #3 Connecticut Shade       15.6%  (customers spend more per stick on Connecticut)
  #4 Habano                  13.0%
  Maduro is the single strongest predictor of velocity at this store.

STRENGTH (balanced qty+revenue):
  #1 Medium-Full   41.7%  ← #1 when weighted by revenue (customers pay more for richer)
  #2 Medium        32.3%
  #3 Mild-Medium   17.7%
  #4 Full           5.4%  (niche)
  #5 Mild           2.9%  (niche)
  Sweet spot: medium-full. Medium is safe. Avoid ultra-full or ultra-mild.

VITOLA — among premium cigars ($8+), balanced qty+revenue:
  #1 Toro             16.0%  ← clear #1 for premium cigars
  #2 Gordo / 60-ring  14.3%
  #3 Corona            9.7%
  #4 Robusto           9.5%
  #5 Churchill         9.1%
  #6 Figurado          8.9%  (customers pay a premium for torpedos/belicosos)
  Small-format cigarillos: 25% balanced but mostly cheap house-brand volume —
    not a good signal for ordering new premium cigar lines.

TOP BRANDS by balanced share (revenue matters as much as units):
  Tier 1 — Core (~5–6% each): Oliva, Espinosa Especial
  Tier 2 — Strong (~4–5%): Padrón, Ashton, Perdomo, Curivari, Espinosa Premium
  Tier 3 — Solid (~2–3%): Rocky Patel, RoMa Craft Tobac, AJ Fernandez, Arturo Fuente
  Tier 4 — Supporting: H. Upmann, Hoyo de Monterrey, La Flor Dominicana, Aladino,
            C.L.E., Acid, Romeo y Julieta, Macanudo, Montecristo
  NOTE: Villiger Braniff is #13 balanced (was #3 by units alone) — pure cheap volume,
    not a demand signal for new premium additions.
  NOT CARRIED / NO DEMAND: Davidoff (near zero), OpusX (zero). Avoid ultra-premium brands.

CUSTOMER SEGMENTS (estimated from balanced data):
  Everyday value smoker   ~30%  Cigarillos / house brand under $8 (less dominant than qty-only suggested)
  Maduro enthusiast       ~35%  $10–16 medium-full maduro — Oliva / Perdomo / Curivari
  Casual premium buyer    ~25%  $12–20 Connecticut or Habano — Ashton / Padrón / Fuente
  Occasional indulger     ~10%  Up to $25 special occasion — Padrón / AJ Fernandez

IDEAL NEW CIGAR for this store:
  Maduro or Connecticut wrapper · Medium-full strength · Toro or Gordo vitola
  $12–16 MSRP · Value-to-mid-premium brand in or adjacent to current portfolio
  (Oliva, Perdomo, AJ Fernandez, Rocky Patel, Espinosa, Padrón tier)
""".strip()


# ── Craziness guidance ────────────────────────────────────────────────────────

def _craziness_guidance(craziness: int) -> str:
    """
    Return a ranking-priority instruction string for the given craziness level (0–10).
    0 = safe/high-fit only.  10 = pure buzz, ignore fit.
    """
    if craziness <= 2:
        return (
            "RANKING PRIORITY — SAFE (craziness 0–2): "
            "Weight ranking: fit 75%, buzz 25%. "
            "Exclude any cigar with fit_score below 55. "
            "We want reliable, proven-profile additions with minimal risk."
        )
    elif craziness <= 4:
        return (
            "RANKING PRIORITY — CONSERVATIVE (craziness 3–4): "
            "Weight ranking: fit 60%, buzz 40%. "
            "Exclude extreme mismatches (fit_score below 35). "
            "A moderately buzzy cigar that fits our profile beats a hugely buzzy mismatch."
        )
    elif craziness <= 6:
        return (
            "RANKING PRIORITY — BALANCED (craziness 5–6): "
            "Weight ranking: fit 50%, buzz 50%. "
            "Include interesting reaches if buzz is very high. "
            "Note mismatches clearly in fit_notes."
        )
    elif craziness <= 8:
        return (
            "RANKING PRIORITY — ADVENTUROUS (craziness 7–8): "
            "Weight ranking: buzz 70%, fit 30%. "
            "Include premium and unusual items even if outside our typical profile. "
            "High buzz justifies the risk. Note fit issues honestly in fit_notes."
        )
    else:
        return (
            "RANKING PRIORITY — WILD (craziness 9–10): "
            "Rank entirely by buzz_score — fit is informational only. "
            "Include everything generating significant online excitement. "
            "We want the most talked-about cigars regardless of whether they "
            "match our store profile."
        )

# ── XLSX styling (matches cigar_researcher.py) ─────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="EBF0FA")
BODY_FONT   = Font(name="Calibri", size=10)

# ── Anthropic web search tool (native, same as cigar_researcher.py) ──────────
# Default caps — override per-call via _make_tool_def()
REPUTATION_MAX_SEARCHES = 6   # per-SKU reputation research (Halfwheel + CA + BMP + Reddit)
BUZZ_MAX_SEARCHES       = 8   # buzz feed refresh (broader sweep, more sources)


def _make_tool_def(max_uses: int) -> dict:
    """Return a web_search tool definition with the given max_uses cap."""
    return {
        "type": "web_search_20250305",
        "name": "web_search",
        "max_uses": max_uses,
    }

# ── system prompts ─────────────────────────────────────────────────────────────
REPUTATION_SYSTEM_PROMPT = """You are a social intelligence analyst for a premium cigar retailer called Smoke Shoppe.

Your job: research the online reputation and community sentiment for a specific cigar, then return a structured JSON report.

SEARCH STRATEGY — execute searches in this order, stopping when you have enough data:
1. Professional reviews:
   • "site:halfwheel.com [brand] [product name]"
   • "[brand] [product name] cigar review score rating"
   • "blind man's puff [brand] [product name]"
   • "cigar aficionado [brand] [product name] rating"
2. Community (only if Reddit/YouTube data is not already provided in the prompt):
   • "site:reddit.com/r/cigars [brand] [product name]"
   • "youtube [brand] [product name] cigar review"

SCORING — all scores are integers 0-100:

quality_score — aggregate of professional reviews:
  Normalize all ratings to 0-100 first:
    • Cigar Aficionado, BMP, Cigar Journal: score is already 0-100
    • Halfwheel qualitative: Buy a Box=97, Buy Several=93, Buy One=86, Try One=77, Pass=55, Don't Buy=30
    • 5-star scales: multiply by 20
    • 4-star scales: multiply by 25
  Weight: Halfwheel 35%, BMP 30%, CA 20%, other sources 15%.
  Use the single found score if only one source.
  null if no professional reviews found.

community_score — community/social sentiment:
  Based on Reddit upvote ratios, tone of discussion, frequency of positive vs. negative mentions,
  YouTube engagement. Scale:
    95+ = beloved classic  |  80-94 = very well regarded  |  65-79 = generally liked
    50-64 = mixed opinions  |  35-49 = disappointing  |  <35 = poorly regarded
  null if no community data found.

value_score — quality relative to price (if MSRP is provided):
  A $10-15 cigar scoring 88 quality ≈ value 94.
  A $20-25 cigar scoring 88 quality ≈ value 83.
  A $35-45 cigar scoring 88 quality ≈ value 70.
  A $50+ cigar scoring 88 quality ≈ value 60.
  Adjust up if widely praised as a deal; adjust down if community considers it overpriced.
  null if quality_score or MSRP unknown.

overall_score = round(0.40 * quality + 0.35 * community + 0.25 * value)
  If value_score is null: round(0.55 * quality + 0.45 * community).
  If community_score is null: round(0.65 * quality + 0.35 * value) or just quality_score.
  null if quality_score is also null.

top_quotes — 2-3 verbatim quotes found in actual search results, pipe-separated, each attributed to its source URL or publication:
  Example: '"Consistently excellent, an easy Buy a Box" — Halfwheel | "Best everyday smoke under $15" — r/cigars'
  HALLUCINATION RULE: Every quote must be text you found verbatim in a search result during THIS session. Never paraphrase, reconstruct, or invent a quote. If you do not have at least one real verbatim quote from a search result, output null. Do not write a plausible-sounding quote from model memory.

source_breakdown — a JSON-encoded string, e.g.:
  '{"halfwheel": {"score": 92, "verdict": "Buy a Box"}, "cigar_aficionado": {"score": 88}, "reddit": {"post_count": 12, "avg_upvote_ratio": 0.95}}'

source_urls — REQUIRED: include at least one URL per non-null score field. If you cannot supply a URL confirming a score, set that score to null instead.

OUTPUT — return ONLY this JSON object, no markdown fences, no extra text:
{
  "overall_score": integer or null,
  "quality_score": integer or null,
  "value_score": integer or null,
  "community_score": integer or null,
  "review_count": integer,
  "reddit_mentions": integer,
  "youtube_videos": integer,
  "top_quotes": string or null,
  "source_breakdown": string,
  "source_urls": [string],
  "notes": string or null
}"""


BUZZ_SYSTEM_PROMPT = """You are a social intelligence analyst for a premium cigar retailer called Smoke Shoppe.

Your job: identify new and upcoming premium cigars generating significant online buzz in 2025-2026.
Focus on RECENT releases — prioritize cigars announced or released in 2025 or 2026.
Do NOT include cigars announced before January 2024.

SEARCH STRATEGY (use your searches efficiently):
1. "site:halfwheel.com 2025 2026 new release" — Halfwheel covers nearly every new release
2. "new cigar releases 2025 2026 announced best"
3. "PCA 2025 2026 new cigars trade show" — Premium Cigar Association announcements
4. "most anticipated cigars 2025 2026 site:reddit.com/r/cigars"
5. Targeted brand searches for high-buzz producers: "Perdomo new 2025", "Oliva new 2025 2026",
   "AJ Fernandez new 2025", "Espinosa new 2025 2026", "Rocky Patel new 2025 2026",
   "My Father 2025 2026", "Liga Privada 2025 2026"
6. "limited edition cigars 2025 2026"
7. "site:halfwheel.com upcoming 2026"

BUZZ SCORE (0-100):
  90+  : must-have of the year, universal excitement
  70-89 : strong buzz, actively discussed
  50-69 : moderate interest
  30-49 : low buzz, niche interest
  <30   : minimal buzz

FIT SCORE (0-100) — how well the cigar matches the store profile provided in the user message:
  90-100 : perfect fit (maduro/habano wrapper, medium-full, toro/gordo, $12-18)
  70-89  : good fit (mostly matches, minor deviations)
  50-69  : moderate fit (some mismatches but workable)
  30-49  : poor fit (price too high, or wrong wrapper/strength profile)
  0-29   : very poor fit (Davidoff/ultra-premium tier, or zero demand profile)
  null   : if no fit profile was provided in the user message

FIT NOTES — brief comma-separated list of fit positives and negatives, e.g.:
  "Maduro ✓, Toro ✓, $16 ✓" or "Habano ✓, $28 — above sweet spot" or "$45 ✗, ultra-premium ✗"
  null if no fit profile was provided.

SENTIMENT:
  "positive" : community is excited, favorable reception
  "mixed"    : divided opinions or concerns about price/availability
  "negative" : generally disappointing reception

RELEASE STATUS:
  "announced" : officially announced, not yet shipping
  "limited"   : shipping but very limited availability
  "released"  : widely available at retail

HALLUCINATION RULE: Every cigar you list must appear in an actual search result you retrieved during THIS session. Do not include cigars from model memory that were not found in a search result. The summary field must describe only what search results actually said — do not invent reception or community reactions. source_urls must contain at least one URL confirming the cigar exists and was announced/released in 2025-2026.

Return ONLY the JSON array — start with [ and end with ], no markdown, no prose:
[
  {
    "name": string,
    "brand": string,
    "parent_company": string or null,
    "announced_date": string or null,
    "release_status": "announced" | "limited" | "released",
    "buzz_score": integer 0-100,
    "fit_score": integer 0-100 or null,
    "reddit_mentions": integer,
    "youtube_videos": integer,
    "sentiment": "positive" | "mixed" | "negative",
    "summary": string,
    "fit_notes": string or null,
    "source_urls": [string]
  },
  ...
]"""


# ── JSON extraction ───────────────────────────────────────────────────────────

def _extract_json_object(text: str) -> dict:
    """Pull a JSON object {...} from a model response (handles fences, prose, truncation)."""
    if not text:
        return {}
    text = text.strip()
    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
        # Try repairing a truncated object
        try:
            return _repair_json(match.group(0))
        except Exception:
            pass
    # Strip ``` fences and try directly
    clean = re.sub(r'^```(?:json)?\s*', '', text, flags=re.MULTILINE)
    clean = re.sub(r'```\s*$', '', clean, flags=re.MULTILINE).strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        pass
    logging.debug("Could not parse JSON object from: %s", text[:200])
    return {}


def _extract_json_array(text: str) -> list:
    """
    Pull a JSON array [...] from a model response.
    Handles: plain JSON, markdown fences, prose before/after the array,
    and truncated responses.
    """
    if not text:
        return []
    text = text.strip()
    # Try the whole text first
    try:
        val = json.loads(text)
        if isinstance(val, list):
            return val
    except json.JSONDecodeError:
        pass
    # Find the LARGEST [...] block in the text (handles prose before/after)
    # Use a greedy match so we capture the full outermost array, not a nested one.
    best: list = []
    for match in re.finditer(r'\[[\s\S]*?\]', text, re.DOTALL):
        candidate = match.group(0)
        # Expand greedily: find the matching close bracket from the opening
        start = match.start()
        depth = 0
        end = start
        for i, ch in enumerate(text[start:], start):
            if ch == '[':
                depth += 1
            elif ch == ']':
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break
        candidate = text[start:end]
        try:
            val = json.loads(candidate)
            if isinstance(val, list) and len(val) > len(best):
                best = val
        except json.JSONDecodeError:
            pass
    if best:
        return best
    # Strip ``` fences and try the inner content directly
    clean = re.sub(r'^```(?:json)?\s*', '', text, flags=re.MULTILINE)
    clean = re.sub(r'```\s*$', '', clean, flags=re.MULTILINE).strip()
    try:
        val = json.loads(clean)
        if isinstance(val, list):
            return val
    except json.JSONDecodeError:
        pass
    logging.debug("Could not parse JSON array from: %s", text[:200])
    return []


def _extract_json_array_from_blocks(content_blocks: list) -> list:
    """
    Search all text blocks in a response for a JSON array.
    Tries each block individually (last first), then falls back to
    the full concatenated text.  Returns the largest array found.
    """
    text_blocks = [b.text for b in content_blocks if hasattr(b, "text") and b.text]
    if not text_blocks:
        return []

    best: list = []
    # Try blocks from last to first (final answer is usually at the end)
    for text in reversed(text_blocks):
        parsed = _extract_json_array(text)
        if len(parsed) > len(best):
            best = parsed

    if best:
        return best

    # Last resort: concatenate everything and try once more
    combined = "\n".join(text_blocks)
    result = _extract_json_array(combined)
    if not result:
        logging.warning(
            "Could not extract JSON array from any of %d text block(s). "
            "First block preview: %s",
            len(text_blocks),
            text_blocks[0][:200] if text_blocks else "(empty)",
        )
    return result


def _repair_json(text: str) -> dict:
    """Best-effort recovery for a JSON object cut off by a token limit."""
    last_comma = text.rfind(',')
    if last_comma != -1:
        text = text[:last_comma]
    depth_curly  = text.count('{') - text.count('}')
    depth_square = text.count('[') - text.count(']')
    if text.rstrip()[-1:] not in ('"', '}', ']', '0123456789'):
        text = text.rstrip().rstrip(',') + '"'
    text += ']' * max(depth_square, 0)
    text += '}' * max(depth_curly, 0)
    return json.loads(text)


# ── cache key helpers ─────────────────────────────────────────────────────────

def _social_cache_key(description: str, brand: str) -> str:
    return f"{str(description).strip().lower()}|{str(brand).strip().lower()}"


def _buzz_cache_key(name: str, brand: str) -> str:
    return f"{str(name).strip().lower()}|{str(brand).strip().lower()}"


# ── main agent class ──────────────────────────────────────────────────────────

class SocialIntelAgent:
    """
    Researches and caches social reputation (Cigar_Social.xlsx)
    and buzz feed (Cigar_Buzz.xlsx) using Claude + optional Reddit/YouTube data.
    """

    def __init__(self, model: str = "claude-sonnet-4-6") -> None:
        self.client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        self.model  = model
        self._ensure_social_file()
        self._ensure_buzz_file()

    # ── XLSX creation / migration ─────────────────────────────────────────────

    def _ensure_social_file(self) -> None:
        if SOCIAL_FILE.exists():
            self._migrate_columns(SOCIAL_FILE, SOCIAL_COLUMNS)
            return
        self._create_xlsx(SOCIAL_FILE, SOCIAL_COLUMNS, "Cigar Social Intel", {
            "Item Number": 14, "Description": 40, "Brand": 22, "Parent Company": 28,
            "Overall Score": 14, "Quality Score": 14, "Value Score": 14, "Community Score": 16,
            "Review Count": 14, "Reddit Mentions": 16, "YouTube Videos": 16,
            "Top Quotes": 70, "Source Breakdown": 50, "Source URLs": 50,
            "Research Date": 16, "Notes": 40,
        })

    def _ensure_buzz_file(self) -> None:
        if BUZZ_FILE.exists():
            self._migrate_columns(BUZZ_FILE, BUZZ_COLUMNS)
            return
        self._create_xlsx(BUZZ_FILE, BUZZ_COLUMNS, "Cigar Buzz Feed", {
            "Name": 40, "Brand": 22, "Parent Company": 28,
            "Announced Date": 18, "Release Status": 14,
            "Buzz Score": 12, "Fit Score": 12,
            "Reddit Mentions": 16, "YouTube Videos": 16,
            "Sentiment": 12, "Summary": 70, "Fit Notes": 40,
            "Source URLs": 50, "Research Date": 16,
        })

    @staticmethod
    def _create_xlsx(path: Path, columns: list[str], sheet_title: str, widths: dict) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(1, col_idx, col_name)
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        wb.save(path)
        logging.info("Created %s", path)

    @staticmethod
    def _migrate_columns(path: Path, columns: list[str]) -> None:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        existing = [cell.value for cell in ws[1]]
        missing = [c for c in columns if c not in existing]
        if not missing:
            return
        logging.info("Migrating %s: adding %s", path.name, missing)
        for col_name in missing:
            col_idx = columns.index(col_name) + 1
            ws.insert_cols(col_idx)
            cell = ws.cell(1, col_idx, col_name)
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        wb.save(path)

    # ── cache loaders ─────────────────────────────────────────────────────────

    def load_social_cache(self) -> dict[str, dict]:
        """Return {cache_key: record} for all rows in Cigar_Social.xlsx."""
        wb = openpyxl.load_workbook(SOCIAL_FILE, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {}
        cache: dict[str, dict] = {}
        for row in rows[1:]:
            record = dict(zip(SOCIAL_COLUMNS, row))
            key = _social_cache_key(
                record.get("Description", ""), record.get("Brand", "")
            )
            cache[key] = record
        return cache

    def load_buzz_cache(self) -> dict[str, dict]:
        """Return {cache_key: record} for all rows in Cigar_Buzz.xlsx."""
        wb = openpyxl.load_workbook(BUZZ_FILE, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {}
        cache: dict[str, dict] = {}
        for row in rows[1:]:
            record = dict(zip(BUZZ_COLUMNS, row))
            key = _buzz_cache_key(record.get("Name", ""), record.get("Brand", ""))
            cache[key] = record
        return cache

    # ── XLSX writers ──────────────────────────────────────────────────────────

    def _save_social(self, result: dict) -> None:
        wb = openpyxl.load_workbook(SOCIAL_FILE)
        ws = wb.active
        key = _social_cache_key(result.get("Description", ""), result.get("Brand", ""))
        target_row = None
        for row in ws.iter_rows(min_row=2):
            desc  = str(row[SOCIAL_COLUMNS.index("Description")].value or "")
            brand = str(row[SOCIAL_COLUMNS.index("Brand")].value or "")
            if _social_cache_key(desc, brand) == key:
                target_row = row[0].row
                break
        if target_row is None:
            target_row = ws.max_row + 1
        is_alt = (target_row % 2 == 0)
        wrap_cols = {"Top Quotes", "Source Breakdown", "Source URLs", "Notes"}
        for col_idx, col_name in enumerate(SOCIAL_COLUMNS, 1):
            cell = ws.cell(target_row, col_idx, result.get(col_name))
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name in wrap_cols))
            if is_alt:
                cell.fill = ALT_FILL
        wb.save(SOCIAL_FILE)

    def _save_buzz(self, result: dict) -> None:
        wb = openpyxl.load_workbook(BUZZ_FILE)
        ws = wb.active
        key = _buzz_cache_key(result.get("Name", ""), result.get("Brand", ""))
        target_row = None
        for row in ws.iter_rows(min_row=2):
            name  = str(row[BUZZ_COLUMNS.index("Name")].value or "")
            brand = str(row[BUZZ_COLUMNS.index("Brand")].value or "")
            if _buzz_cache_key(name, brand) == key:
                target_row = row[0].row
                break
        if target_row is None:
            target_row = ws.max_row + 1
        is_alt = (target_row % 2 == 0)
        wrap_cols = {"Summary", "Source URLs"}
        for col_idx, col_name in enumerate(BUZZ_COLUMNS, 1):
            cell = ws.cell(target_row, col_idx, result.get(col_name))
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name in wrap_cols))
            if is_alt:
                cell.fill = ALT_FILL
        wb.save(BUZZ_FILE)

    # ── MSRP cross-reference ──────────────────────────────────────────────────

    def _get_msrp(self, description: str, brand: str) -> float | None:
        """Look up MSRP from Cigar_Research.xlsx if available."""
        if not RESEARCH_FILE.exists():
            return None
        try:
            wb = openpyxl.load_workbook(RESEARCH_FILE, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            if "MSRP" not in headers or "Description" not in headers:
                return None
            msrp_idx = headers.index("MSRP")
            desc_idx = headers.index("Description")
            brand_idx = headers.index("Brand") if "Brand" in headers else None
            needle = _social_cache_key(description, brand)
            for row in ws.iter_rows(min_row=2, values_only=True):
                desc  = str(row[desc_idx] or "")
                brd   = str(row[brand_idx] or "") if brand_idx is not None else ""
                if _social_cache_key(desc, brd) == needle:
                    val = row[msrp_idx]
                    return float(val) if val else None
        except Exception:
            pass
        return None

    # ── agent loops ───────────────────────────────────────────────────────────

    def research_reputation(
        self,
        description: str,
        brand: str,
        item_number: str | None = None,
        parent_company: str | None = None,
        max_searches: int = REPUTATION_MAX_SEARCHES,
    ) -> dict[str, Any]:
        """
        Research social reputation for one cigar SKU.
        Gathers Reddit/YouTube data first (if configured), then runs Claude.

        max_searches — cap on Anthropic web_search calls (default: REPUTATION_MAX_SEARCHES).
                       Lower values reduce cost; higher values improve coverage.
        """
        logging.info("Social research: %s (%s)", description, brand)

        query = f"{brand} {description}".strip()

        # Pre-gather optional enrichment
        reddit_posts, reddit_warning = reddit_search(query)
        youtube_videos, youtube_warning = youtube_search(query)
        msrp = self._get_msrp(description, brand)

        # Build user message
        msrp_str = f"${msrp:.2f}" if msrp else "unknown"
        reddit_block   = reddit_format(reddit_posts, reddit_warning)
        youtube_block  = youtube_format(youtube_videos, youtube_warning)

        user_message = (
            f"Research the social reputation and community sentiment for this cigar:\n"
            f"  Description : {description}\n"
            f"  Brand       : {brand}\n"
            f"  MSRP        : {msrp_str}\n\n"
            f"--- Reddit Community Data ---\n{reddit_block}\n\n"
            f"--- YouTube Data ---\n{youtube_block}\n\n"
            f"Now search for professional reviews (Halfwheel, Cigar Aficionado, Blind Man's Puff, etc.) "
            f"and compute all reputation scores as specified."
        )

        messages = [{"role": "user", "content": user_message}]
        raw_json: dict = {}

        for _turn in range(max_searches + 4):
            response = _create_with_backoff(
                self.client,
                model=self.model,
                max_tokens=4096,
                system=REPUTATION_SYSTEM_PROMPT,
                tools=[_make_tool_def(max_searches)],
                messages=messages,
            )
            messages.append({"role": "assistant", "content": response.content})

            if response.stop_reason == "end_turn":
                # Search all text blocks (model sometimes puts prose after JSON)
                text_blocks = [b.text for b in response.content if hasattr(b, "text") and b.text]
                for text in reversed(text_blocks):
                    candidate = _extract_json_object(text)
                    if candidate:
                        raw_json = candidate
                        break
                else:
                    # Fallback: try concatenated text
                    combined = "\n".join(text_blocks)
                    raw_json = _extract_json_object(combined)
                    if not raw_json:
                        logging.warning(
                            "Could not extract JSON object for %s (%s) from %d block(s).",
                            description, brand, len(text_blocks),
                        )
                break

            # web_search_20250305 is a server-side tool: results are embedded in
            # the assistant message, but the model may need multiple rounds to finish.
            # The API requires user/assistant alternation — add a continuation
            # user turn so the next call doesn't fail with "assistant prefill" error.
            if response.stop_reason in ("tool_use", "max_tokens"):
                messages.append({
                    "role": "user",
                    "content": (
                        "Continue. When finished, output ONLY the raw JSON object — "
                        "no prose, no markdown fences, no explanation. "
                        "Start your response with { and end with }."
                    ),
                })
                continue

            break  # unexpected stop reason

        # Compile notes about missing enrichment
        notes_parts = []
        if reddit_warning:
            notes_parts.append(f"Reddit: {reddit_warning}")
        if youtube_warning:
            notes_parts.append(f"YouTube: {youtube_warning}")
        if raw_json.get("notes"):
            notes_parts.append(raw_json["notes"])

        source_urls = raw_json.get("source_urls", [])
        if isinstance(source_urls, list):
            source_urls = "\n".join(source_urls)

        result = {
            "Item Number":      item_number,
            "Description":      description,
            "Brand":            brand,
            "Parent Company":   parent_company,
            "Overall Score":    raw_json.get("overall_score"),
            "Quality Score":    raw_json.get("quality_score"),
            "Value Score":      raw_json.get("value_score"),
            "Community Score":  raw_json.get("community_score"),
            "Review Count":     raw_json.get("review_count", 0),
            "Reddit Mentions":  raw_json.get("reddit_mentions", len(reddit_posts)),
            "YouTube Videos":   raw_json.get("youtube_videos", len(youtube_videos)),
            "Top Quotes":       raw_json.get("top_quotes"),
            "Source Breakdown": raw_json.get("source_breakdown"),
            "Source URLs":      source_urls,
            "Research Date":    str(date.today()),
            "Notes":            " | ".join(notes_parts) if notes_parts else None,
        }

        self._save_social(result)
        return result

    def refresh_buzz_feed(
        self,
        max_searches: int = BUZZ_MAX_SEARCHES,
        target_count: int = 25,
        fit_profile: str | None = DEFAULT_FIT_PROFILE,
        craziness: int = 5,
        since_months: int = 3,
    ) -> list[dict[str, Any]]:
        """
        Run a Claude web search pass to find new/upcoming cigars generating buzz.
        Upserts results into Cigar_Buzz.xlsx.

        max_searches  — cap on web_search calls (default: BUZZ_MAX_SEARCHES).
        target_count  — how many NEW (uncached) buzz cigars to find (default: 15).
        fit_profile   — store profile text used to compute fit_score per cigar.
                        Defaults to DEFAULT_FIT_PROFILE. Pass None to skip fit scoring.
        craziness     — 0-10 scale controlling fit vs. buzz tradeoff in ranking.
                        0 = safe/high-fit only. 10 = pure buzz, ignore fit.
        since_months  — only find cigars announced/released within the last N months.
                        Default 3. Pass 0 to disable the date filter.
        """
        craziness = max(0, min(10, craziness))

        # ── date cutoff ───────────────────────────────────────────────────────
        if since_months > 0:
            since_date = date.today() - timedelta(days=since_months * 30)
            since_str  = since_date.strftime("%B %-d, %Y")   # e.g. "March 7, 2026"
            date_instruction = (
                f"DATE FILTER: Only include cigars that were first announced or released "
                f"on or after {since_str} (last {since_months} months). "
                f"Do NOT include anything announced before that date.\n"
            )
        else:
            since_str = None
            date_instruction = ""

        # ── skip already-cached cigars ────────────────────────────────────────
        existing = self.load_buzz_cache()
        cached_names = sorted(
            f"{v.get('Name', '')} ({v.get('Brand', '')})"
            for v in existing.values()
            if v.get("Name")
        )
        if cached_names:
            skip_section = (
                "ALREADY IN DATABASE — do not include any of these, "
                "even if they were announced recently:\n"
                + "\n".join(f"  - {n}" for n in cached_names)
                + "\n"
            )
        else:
            skip_section = ""

        logging.info(
            "Refreshing buzz feed (max_searches=%d, target=%d, craziness=%d, since=%s, skip=%d cached)…",
            max_searches, target_count, craziness,
            since_str or "no date filter", len(cached_names),
        )

        # ── Reddit enrichment (scoped to the same window) ─────────────────────
        reddit_query = (
            f"new cigar release {since_date.year}" if since_months > 0
            else "new cigar release 2025 2026"
        )
        reddit_posts, reddit_warning = reddit_search(reddit_query, time_filter="month")
        reddit_block = reddit_format(reddit_posts, reddit_warning)

        # ── fit profile section ───────────────────────────────────────────────
        if fit_profile:
            fit_section = (
                f"\n--- STORE FIT PROFILE ---\n{fit_profile}\n\n"
                f"--- RANKING GUIDANCE ---\n{_craziness_guidance(craziness)}\n"
            )
        else:
            fit_section = "\nNo fit profile provided — set fit_score and fit_notes to null for all items.\n"

        user_message = (
            f"Find {target_count} NEW premium cigars that are generating online buzz and that "
            f"are NOT yet in our database.\n\n"
            f"{date_instruction}"
            f"{skip_section}\n"
            f"--- Recent Reddit r/cigars posts about new releases ---\n{reddit_block}\n"
            f"{fit_section}\n"
            "Search the web for new releases, PCA announcements, Halfwheel coverage, "
            f"and other buzz signals. Score each cigar against the fit profile above. "
            f"Return a JSON array of exactly {target_count} items "
            f"(you have {max_searches} web searches — use them efficiently). "
            f"Start your response with [ and end with ]."
        )

        messages = [{"role": "user", "content": user_message}]
        buzz_items: list[dict] = []

        for _turn in range(max_searches + 4):  # turns = searches + headroom for reasoning
            response = _create_with_backoff(
                self.client,
                model=self.model,
                max_tokens=8192,   # buzz feed JSON for 15+ cigars can be large
                system=BUZZ_SYSTEM_PROMPT,
                tools=[_make_tool_def(max_searches)],
                messages=messages,
            )
            messages.append({"role": "assistant", "content": response.content})

            if response.stop_reason == "end_turn":
                buzz_items = _extract_json_array_from_blocks(response.content)
                break

            # web_search_20250305 is a server-side tool: results are embedded in
            # the assistant message, but the model may need multiple rounds to finish.
            # The API requires user/assistant alternation — add a continuation
            # user turn so the next call doesn't fail with "assistant prefill" error.
            if response.stop_reason in ("tool_use", "max_tokens"):
                messages.append({
                    "role": "user",
                    "content": (
                        "Continue. When you have finished all searches, output ONLY "
                        "the raw JSON array — no prose, no markdown fences, no explanation. "
                        "Start your response with [ and end with ]."
                    ),
                })
                continue

            break  # unexpected stop reason

        results = []
        today = str(date.today())
        for item in buzz_items:
            source_urls = item.get("source_urls", [])
            if isinstance(source_urls, list):
                source_urls = "\n".join(source_urls)
            record = {
                "Name":           item.get("name", ""),
                "Brand":          item.get("brand", ""),
                "Parent Company": item.get("parent_company"),
                "Announced Date": item.get("announced_date"),
                "Release Status": item.get("release_status"),
                "Buzz Score":     item.get("buzz_score"),
                "Fit Score":      item.get("fit_score"),
                "Reddit Mentions":item.get("reddit_mentions", 0),
                "YouTube Videos": item.get("youtube_videos", 0),
                "Sentiment":      item.get("sentiment"),
                "Summary":        item.get("summary"),
                "Fit Notes":      item.get("fit_notes"),
                "Source URLs":    source_urls,
                "Research Date":  today,
            }
            if record["Name"]:
                self._save_buzz(record)
                results.append(record)

        return results

    # ── batch reputation research ─────────────────────────────────────────────

    def batch_research(
        self,
        limit: int | None = None,
        force: bool = False,
        rate_limit_s: float = 15.0,
        category: str = "Cigars",
        since: str | None = None,
        sort_by_sales: bool = False,
        exclude_brands: tuple[str, ...] = ("Smoke Shoppe",),
        exclude_parent_companies: tuple[str, ...] = (),
    ) -> list[dict]:
        """
        Batch-research inventory cigars, optionally filtered and sorted by sales.

        since         – timeframe string: "2025", "last 6 months", "last 90 days",
                        "Q2 2024", or "2024-01 to 2024-06".  Filters to items
                        that had at least one sale within that window.
        sort_by_sales – if True, research highest-selling items first, ranked by
                        a balanced qty+revenue score within the ``since`` window
                        (or all-time if since is None).
        exclude_brands          – brand names to skip (default: Smoke Shoppe house brands).
        exclude_parent_companies – parent companies to skip (default: Smoke Shoppe).
                                   Catches any sub-brand under the house parent.
        """
        from tools.inventory_tool import run_inventory_sql_df
        conditions = []
        if category:
            conditions.append(f"Category = '{category}'")
        if exclude_brands:
            brands_sql = ", ".join(f"'{b.replace(chr(39), chr(39)*2)}'" for b in exclude_brands)
            conditions.append(f"Brand NOT IN ({brands_sql})")
        if exclude_parent_companies:
            pcs_sql = ", ".join(f"'{p.replace(chr(39), chr(39)*2)}'" for p in exclude_parent_companies)
            conditions.append(f'"Parent Company" NOT IN ({pcs_sql})')
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        inv = run_inventory_sql_df(
            f'SELECT "Item Number", Description, Brand, "Parent Company" '
            f"FROM inventory {where}",
            file_path=str(INVENTORY_FILE),
        )

        if since or sort_by_sales:
            inv = _apply_sales_filter(inv, since=since, sort_by_sales=sort_by_sales)

        cache = {} if force else self.load_social_cache()
        results = []
        skipped = count = 0

        for _, row in inv.iterrows():
            description    = str(row.get("Description", "") or "")
            brand          = str(row.get("Brand", "") or "")
            item_number    = str(row.get("Item Number", "") or "")
            parent_company = str(row.get("Parent Company", "") or "")

            if not description or description == "nan":
                continue

            key = _social_cache_key(description, brand)
            if key in cache:
                skipped += 1
                continue

            units   = row.get("_units_sold", "")
            revenue = row.get("_revenue", "")
            if units != "" and revenue != "":
                sales_str = f"  {int(units)} units / ${revenue:,.0f} rev"
            elif units != "":
                sales_str = f"  {int(units)} units sold"
            else:
                sales_str = ""
            try:
                result = self.research_reputation(description, brand, item_number, parent_company)
                results.append(result)
                count += 1
                score = result.get("Overall Score")
                print(f"[{count}] ✓ {description[:50]}  (overall: {score}){sales_str}", flush=True)
            except Exception as exc:
                logging.error("Error on %s: %s", description, exc)
                print(f"[{count}] ✗ {description[:50]} — {exc}", flush=True)

            if limit and count >= limit:
                break

            time.sleep(rate_limit_s)

        print(f"\nDone. Researched: {count}  |  Skipped (cached): {skipped}")
        return results


# ── singleton ─────────────────────────────────────────────────────────────────

_agent: SocialIntelAgent | None = None


def _get_agent() -> SocialIntelAgent:
    global _agent
    if _agent is None:
        _agent = SocialIntelAgent()
    return _agent


# ── public API (used by MCP server and other agents) ─────────────────────────

def lookup_social(
    description: str,
    brand: str = "",
    item_number: str = "",
    parent_company: str = "",
    use_cache: bool = True,
) -> dict[str, Any]:
    """
    Return social reputation for one cigar.
    Hits cache first; runs live research if not yet cached (or use_cache=False).
    """
    agent = _get_agent()
    if use_cache:
        cache = agent.load_social_cache()
        key = _social_cache_key(description, brand)
        if key in cache:
            return cache[key]
    return agent.research_reputation(description, brand, item_number, parent_company)


def get_all_social() -> list[dict]:
    """Return every row in Cigar_Social.xlsx as a list of dicts."""
    return list(_get_agent().load_social_cache().values())


def get_buzz_feed(
    refresh: bool = False,
    max_searches: int = BUZZ_MAX_SEARCHES,
    target_count: int = 25,
    fit_profile: str | None = DEFAULT_FIT_PROFILE,
    craziness: int = 5,
    since_months: int = 3,
) -> list[dict]:
    """
    Return the current buzz feed (Cigar_Buzz.xlsx).
    If refresh=True, runs a new web search pass first.

    max_searches  — web search cap for the refresh pass (default: BUZZ_MAX_SEARCHES).
    target_count  — number of NEW (uncached) buzz cigars to find (default: 15).
    fit_profile   — store profile used to score fit. Defaults to DEFAULT_FIT_PROFILE.
    craziness     — 0-10: 0=safe/high-fit, 10=pure buzz. Default: 5 (balanced).
    since_months  — only find cigars from the last N months (default 3, used when refresh=True).
    """
    agent = _get_agent()
    if refresh:
        agent.refresh_buzz_feed(
            max_searches=max_searches,
            target_count=target_count,
            fit_profile=fit_profile,
            craziness=craziness,
            since_months=since_months,
        )
    return list(agent.load_buzz_cache().values())


def social_status() -> dict:
    """Return coverage summary for both caches."""
    from tools.inventory_tool import run_inventory_sql
    agent = _get_agent()
    total_cigars = run_inventory_sql(
        "SELECT COUNT(*) FROM inventory WHERE Category = 'Cigars'",
        file_path=str(INVENTORY_FILE),
    )[0][0]
    social_cache = agent.load_social_cache()
    buzz_cache   = agent.load_buzz_cache()

    from tools.reddit_tool import is_available as reddit_ok, availability_note as reddit_note
    from tools.youtube_tool import is_available as youtube_ok, availability_note as youtube_note

    return {
        "total_cigar_skus":     total_cigars,
        "social_researched":    len(social_cache),
        "social_remaining":     max(0, total_cigars - len(social_cache)),
        "buzz_items":           len(buzz_cache),
        "reddit_configured":    reddit_ok(),
        "reddit_note":          reddit_note() or "OK",
        "youtube_configured":   youtube_ok(),
        "youtube_note":         youtube_note() or "OK",
        "social_file":          str(SOCIAL_FILE),
        "buzz_file":            str(BUZZ_FILE),
    }


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    logging.basicConfig(level=logging.WARNING, format="%(levelname)s: %(message)s")

    parser = argparse.ArgumentParser(description="Smoke Shoppe Social Intelligence Agent")
    parser.add_argument("description", nargs="?", help="Cigar description to research")
    parser.add_argument("brand",       nargs="?", default="", help="Cigar brand")
    parser.add_argument("--batch",        action="store_true", help="Batch-research all uncached inventory cigars")
    parser.add_argument("--since",        type=str,   default=None,
                        help="Only research cigars sold within this window. "
                             "Examples: '2025', 'last 6 months', 'last 90 days', 'Q2 2024', '2024-01 to 2024-06'")
    parser.add_argument("--top",          action="store_true",
                        help="Sort by balanced qty+revenue score (descending) within --since window, or all-time if omitted")
    parser.add_argument("--buzz",         action="store_true", help="Refresh the buzz feed (new/upcoming cigars)")
    parser.add_argument("--limit",        type=int,   default=None,              help="Max items in batch mode")
    parser.add_argument("--force",        action="store_true",                   help="Re-research even if cached")
    parser.add_argument("--delay",        type=float, default=15.0,              help="Seconds between batch requests (default 15)")
    parser.add_argument("--max-searches", type=int,   default=None,
                        help=f"Cap on web searches per agent call. "
                             f"Buzz default: {BUZZ_MAX_SEARCHES}, reputation default: {REPUTATION_MAX_SEARCHES}. "
                             f"Lower = cheaper/faster, higher = more thorough.")
    parser.add_argument("--target",        type=int,   default=25,
                        help="Number of NEW (uncached) buzz cigars to find (default: 15, used with --buzz)")
    parser.add_argument("--since-months", type=int,   default=3,
                        help="Only find cigars announced/released in the last N months (default: 3). "
                             "Pass 0 to disable the date filter. Used with --buzz.")
    parser.add_argument("--craziness",    type=int,   default=5,
                        help="0-10 scale: 0=safe/high-fit only, 10=pure buzz ignore fit. Default: 5 (balanced).")
    parser.add_argument("--no-fit",       action="store_true",
                        help="Disable fit scoring (no fit profile sent to Claude). Faster, no fit_score output.")
    parser.add_argument("--status",       action="store_true",                   help="Show cache and API status, then exit")
    args = parser.parse_args()

    if args.status:
        s = social_status()
        print(f"Inventory cigar SKUs   : {s['total_cigar_skus']}")
        print(f"Social cache           : {s['social_researched']} researched, {s['social_remaining']} remaining")
        print(f"Buzz feed items        : {s['buzz_items']}")
        print(f"Reddit configured      : {'✓' if s['reddit_configured'] else '✗'}  {s['reddit_note']}")
        print(f"YouTube configured     : {'✓' if s['youtube_configured'] else '✗'}  {s['youtube_note']}")
        print(f"Social file            : {s['social_file']}")
        print(f"Buzz file              : {s['buzz_file']}")
        sys.exit(0)

    if args.buzz:
        max_s         = args.max_searches if args.max_searches is not None else BUZZ_MAX_SEARCHES
        fit_prof      = None if args.no_fit else DEFAULT_FIT_PROFILE
        craziness     = max(0, min(10, args.craziness))
        since_months  = max(0, args.since_months)
        items = get_buzz_feed(
            refresh=True,
            max_searches=max_s,
            target_count=args.target,
            fit_profile=fit_prof,
            craziness=craziness,
            since_months=since_months,
        )
        fit_label    = f"  craziness={craziness}" if fit_prof else "  (no fit scoring)"
        since_label  = f"  last {since_months}mo" if since_months > 0 else "  (no date filter)"
        print(f"\nBuzz feed refreshed — {len(items)} total items in cache  (≤{max_s} searches{fit_label}{since_label})")
        print(f"{'Buzz':>4}  {'Fit':>4}  {'Sentiment':<10}  Name (Brand)")
        print("─" * 70)
        for item in sorted(items, key=lambda x: -(x.get("Buzz Score") or 0))[:15]:
            buzz = item.get("Buzz Score") or 0
            fit  = item.get("Fit Score")
            fit_str = f"{fit:>4}" if fit is not None else "   —"
            print(f"  {buzz:>3}  {fit_str}  {str(item.get('Sentiment','')):<10}  {item.get('Name','')} ({item.get('Brand','')})")
        sys.exit(0)

    if args.batch:
        _get_agent().batch_research(
            limit=args.limit,
            force=args.force,
            rate_limit_s=args.delay,
            since=args.since,
            sort_by_sales=args.top,
        )
        sys.exit(0)

    if args.description:
        result = lookup_social(args.description, args.brand, use_cache=not args.force)
        print(json.dumps(result, indent=2, default=str))
    else:
        parser.print_help()
