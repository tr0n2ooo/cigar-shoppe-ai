"""
inventory_verifier.py
---------------------
Produces a skeptic-adjusted copy of Smoke_Shoppe_Inventory.xlsx called
Smoke_Shoppe_Inventory_Verified.xlsx by applying two conservative rules:

  Rule 1 — Zero-out never-sold items
    If an inventory item's Item Number never appears in the transactions file,
    its On Hand quantity is set to 0.  If inventory data was entered but no sale
    was ever recorded, we treat the stock figure as unreliable.

  Rule 2 — Clamp negative quantities
    Negative On Hand values are set to 0 (can't physically have negative stock).

A "Verification Notes" column is added to explain each adjustment.

Usage:
    python inventory_verifier.py
    python inventory_verifier.py --inv  data/Smoke_Shoppe_Inventory.xlsx
    python inventory_verifier.py --tx   data/Smoke_Shoppe_Transactions.xlsx
    python inventory_verifier.py --out  data/My_Output.xlsx
    python inventory_verifier.py --summary   (print stats only, don't write file)
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

DATA_DIR   = Path(__file__).parent / "data"
INV_FILE   = DATA_DIR / "Smoke_Shoppe_Inventory.xlsx"
TX_FILE    = DATA_DIR / "Smoke_Shoppe_Transactions.xlsx"
OUT_FILE   = DATA_DIR / "Smoke_Shoppe_Inventory_Verified.xlsx"

# ── styling (matches the rest of the project) ─────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")
BODY_FONT    = Font(name="Calibri", size=10)
CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")   # soft yellow for changed rows
CHANGED_FONT = Font(name="Calibri", size=10, bold=True)

NOTE_COL_NAME = "Verification Notes"


def build_verified_inventory(
    inv_path: str = str(INV_FILE),
    tx_path:  str = str(TX_FILE),
) -> tuple[pd.DataFrame, dict]:
    """
    Load both files, apply the two rules, and return:
      (verified_df, stats)

    verified_df has all original columns plus NOTE_COL_NAME.
    stats is a summary dict for reporting.
    """
    # ── load ──────────────────────────────────────────────────────────────────
    inv = pd.read_excel(inv_path, header=1)
    tx  = pd.read_excel(tx_path)

    inv = inv.copy()

    # Normalise Item Number to string for reliable set lookup
    inv["_item_key"] = inv["Item Number"].fillna("").astype(str).str.strip()
    sold_keys = set(tx["Item Number"].dropna().astype(str).str.strip())

    # ── apply rules ───────────────────────────────────────────────────────────
    notes: list[str] = []
    original_on_hand = inv["On Hand"].copy()

    for idx, row in inv.iterrows():
        key      = row["_item_key"]
        on_hand  = row["On Hand"]
        parts    = []

        never_sold = key not in sold_keys
        negative   = isinstance(on_hand, (int, float)) and on_hand < 0

        if never_sold and negative:
            inv.at[idx, "On Hand"] = 0
            parts.append(f"zeroed: never sold + was negative ({int(on_hand)})")
        elif never_sold:
            inv.at[idx, "On Hand"] = 0
            parts.append("zeroed: never appears in sales data")
        elif negative:
            inv.at[idx, "On Hand"] = 0
            parts.append(f"clamped: negative quantity ({int(on_hand)} → 0)")

        notes.append("; ".join(parts))

    inv[NOTE_COL_NAME] = notes
    inv = inv.drop(columns=["_item_key"])

    # ── stats ─────────────────────────────────────────────────────────────────
    changed_mask   = inv[NOTE_COL_NAME] != ""
    never_sold_ct  = inv[NOTE_COL_NAME].str.contains("never sold", na=False).sum()
    negative_ct    = inv[NOTE_COL_NAME].str.contains("clamped", na=False).sum()
    both_ct        = inv[NOTE_COL_NAME].str.contains(r"never sold \+ was negative", na=False).sum()

    stats = {
        "total_rows":        len(inv),
        "total_changed":     int(changed_mask.sum()),
        "zeroed_never_sold": int(never_sold_ct),
        "clamped_negative":  int(negative_ct),
        "both_rules":        int(both_ct),
        "unchanged":         int(len(inv) - changed_mask.sum()),
    }
    return inv, stats


def write_verified_xlsx(df: pd.DataFrame, out_path: str = str(OUT_FILE)) -> None:
    """Write the verified DataFrame to a formatted xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verified Inventory"

    cols = list(df.columns)

    # ── row 1: blank title row (matches original file layout — all readers use header=1) ──
    ws.row_dimensions[1].height = 18

    # ── row 2: header row ─────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(2, c_idx, col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    # ── data rows (start at row 3) ────────────────────────────────────────────
    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        is_alt     = (r_idx % 2 == 1)   # row 3 = first data row = unshaded; row 4 = alt
        is_changed = bool(row[NOTE_COL_NAME])

        for c_idx, col_name in enumerate(cols, 1):
            val  = row[col_name]
            # Convert NaN / NaT to None so openpyxl writes blank cells
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None

            cell = ws.cell(r_idx, c_idx, val)

            if is_changed:
                cell.fill = CHANGED_FILL
                cell.font = CHANGED_FONT
            elif is_alt:
                cell.fill = ALT_FILL
                cell.font = BODY_FONT
            else:
                cell.font = BODY_FONT

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_name == NOTE_COL_NAME),
            )

    # ── column widths ─────────────────────────────────────────────────────────
    fixed_widths = {
        "Category": 14, "Description": 42, "Brand": 22, "Parent Company": 28,
        "Item Number": 18, "UPC": 18, "On Hand": 10, "Allocated": 10,
        "Reorder Quantity": 14, "Cost": 10, "Average Cost": 14, "On Order": 10,
        "Minimum Level": 14, "Maximum Level": 14, "Retail Price": 12,
        "% Mark Up": 10, "VAT Category": 14, "VAT Amount": 10,
        "Selling Price": 14, NOTE_COL_NAME: 45,
    }
    for c_idx, col_name in enumerate(cols, 1):
        width = fixed_widths.get(col_name, 14)
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    wb.save(out_path)


def print_summary(stats: dict, out_path: str | None = None) -> None:
    print()
    print("=" * 60)
    print("  INVENTORY VERIFICATION SUMMARY")
    print("=" * 60)
    print(f"  Total inventory rows   : {stats['total_rows']:>6,}")
    print(f"  Unchanged              : {stats['unchanged']:>6,}")
    print(f"  ─────────────────────────────")
    print(f"  Zeroed (never sold)    : {stats['zeroed_never_sold']:>6,}  "
          f"(Rule 1: no sales history)")
    print(f"  Clamped (negative→0)   : {stats['clamped_negative']:>6,}  "
          f"(Rule 2: can't have negative stock)")
    print(f"  Both rules applied     : {stats['both_rules']:>6,}")
    print(f"  ─────────────────────────────")
    print(f"  Total rows adjusted    : {stats['total_changed']:>6,}")
    if out_path:
        print(f"\n  Output: {out_path}")
    print("=" * 60)
    print()


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a skeptic-adjusted copy of the inventory file."
    )
    parser.add_argument("--inv",     default=str(INV_FILE),
                        help="Path to Smoke_Shoppe_Inventory.xlsx")
    parser.add_argument("--tx",      default=str(TX_FILE),
                        help="Path to Smoke_Shoppe_Transactions.xlsx")
    parser.add_argument("--out",     default=str(OUT_FILE),
                        help="Output path (default: data/Smoke_Shoppe_Inventory_Verified.xlsx)")
    parser.add_argument("--summary", action="store_true",
                        help="Print summary statistics only — do not write a file")
    args = parser.parse_args()

    print("Loading inventory and transactions…", flush=True)
    df, stats = build_verified_inventory(inv_path=args.inv, tx_path=args.tx)

    print_summary(stats, out_path=None if args.summary else args.out)

    if not args.summary:
        print(f"Writing {args.out}…", flush=True)
        write_verified_xlsx(df, out_path=args.out)
        print("Done.")


if __name__ == "__main__":
    main()
