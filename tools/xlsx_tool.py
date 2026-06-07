"""
tools/xlsx_tool.py
------------------
Reads and queries an XLSX file from the local filesystem.
Exposes four actions:
  - list_sheets   : return all sheet names
  - describe_sheet: return column headers and row count for a sheet
  - read_sheet    : return rows (up to max_rows) as dicts
  - query_data    : return rows whose any cell contains a keyword
"""

import json
from pathlib import Path
from typing import Optional

import openpyxl


class XlsxReaderTool:
    """Read and query an Excel (.xlsx) file."""

    name = "xlsx_reader"
    description = (
        "Read and query an Excel (.xlsx) file. "
        "Use action='list_sheets' to discover sheet names, "
        "action='describe_sheet' to see headers and row count, "
        "action='read_sheet' to retrieve rows, and "
        "action='query_data' to search for specific values."
    )

    def __init__(self, file_path: str):
        self.file_path = str(Path(file_path).resolve())

    def run(
        self,
        action: str,
        sheet_name: Optional[str] = None,
        keyword: Optional[str] = None,
        max_rows: int = 100,
    ) -> str:
        return self._run(action, sheet_name, keyword, max_rows)

    def _load_workbook(self) -> openpyxl.Workbook:
        path = Path(self.file_path)
        if not path.exists():
            raise FileNotFoundError(f"XLSX file not found: {self.file_path}")
        return openpyxl.load_workbook(path, data_only=True, read_only=True)

    def _get_sheet(self, wb: openpyxl.Workbook, sheet_name: Optional[str]):
        if sheet_name and sheet_name in wb.sheetnames:
            return wb[sheet_name]
        return wb.active

    @staticmethod
    def _rows_to_dicts(sheet, max_rows: int) -> list[dict]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        return [dict(zip(headers, row)) for row in rows[1 : max_rows + 1]]

    def _run(
        self,
        action: str,
        sheet_name: Optional[str] = None,
        keyword: Optional[str] = None,
        max_rows: int = 100,
    ) -> str:
        try:
            wb = self._load_workbook()

            if action == "list_sheets":
                return json.dumps({"sheets": wb.sheetnames})

            sheet = self._get_sheet(wb, sheet_name)
            actual_name = sheet.title

            if action == "describe_sheet":
                rows = list(sheet.iter_rows(values_only=True))
                headers = list(rows[0]) if rows else []
                return json.dumps({
                    "sheet": actual_name,
                    "headers": headers,
                    "data_row_count": len(rows) - 1 if len(rows) > 1 else 0,
                })

            if action == "read_sheet":
                return json.dumps(
                    {"sheet": actual_name, "rows": self._rows_to_dicts(sheet, max_rows)},
                    default=str,
                )

            if action == "query_data":
                if not keyword:
                    return json.dumps({"error": "keyword is required for query_data"})
                kw = keyword.lower()
                all_rows = self._rows_to_dicts(sheet, 50_000)
                matches = [
                    r for r in all_rows
                    if any(kw in str(v).lower() for v in r.values() if v is not None)
                ]
                return json.dumps({
                    "sheet": actual_name,
                    "keyword": keyword,
                    "match_count": len(matches),
                    "rows": matches[:max_rows],
                }, default=str)

            return json.dumps({"error": f"Unknown action: {action}"})

        except Exception as exc:
            return json.dumps({"error": str(exc)})
