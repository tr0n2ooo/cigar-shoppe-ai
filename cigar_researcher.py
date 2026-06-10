"""
cigar_researcher.py
-------------------
Standalone cigar research agent powered by Claude + web search.

For each cigar it finds: wrapper, binder, filler, country of origin,
factory, strength, size, flavor notes (manufacturer copy), MSRP/MAP.
Results are cached in data/Cigar_Research.xlsx so each SKU is only
researched once.

Web search is handled natively by the Anthropic API (web_search_20250305) —
no extra API keys required. Searches are billed at $10/1,000 through your
existing Anthropic account.

Usage:
  # Research a single cigar (prints JSON)
  python cigar_researcher.py "Perdomo BBA Mad. Churchill" "Perdomo"

  # Batch-research all uncached inventory cigars
  python cigar_researcher.py --batch

  # Limit batch to first N uncached items (for testing)
  python cigar_researcher.py --batch --limit 10

  # Re-research items even if cached
  python cigar_researcher.py --batch --force
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import threading
import time
from datetime import date
from pathlib import Path
from typing import Any

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

# ── paths ────────────────────────────────────────────────────────────────────
DATA_DIR       = Path(__file__).parent / "data"
RESEARCH_FILE  = DATA_DIR / "Cigar_Research.xlsx"
INVENTORY_FILE = DATA_DIR / "Smoke_Shoppe_Inventory_Verified.xlsx"

# ── Line-key deduplication ────────────────────────────────────────────────────
# Words that appear at the end of a description to denote size/shape but are NOT
# part of the blend line name.  We strip these from the right to get a "line key"
# so that siblings (same blend, different vitola) share a cache entry for blend fields.
_VITOLA_WORDS: frozenset[str] = frozenset({
    # classic vitola names
    'robusto', 'toro', 'churchill', 'corona', 'lancero', 'gordo', 'belicoso',
    'torpedo', 'panetela', 'lonsdale', 'magnum', 'figurado', 'perfecto',
    'pyramid', 'piramide', 'rothschild', 'presidente', 'cetro', 'epicure',
    'gigante', 'canonazo', 'hermosos', 'monarcas', 'imperiales', 'cervantes',
    'nobles', 'artistas', 'angeles', 'regios', 'noellas', 'cazadores',
    # size qualifiers
    'gran', 'grande', 'fino', 'largo', 'petit', 'double', 'triple',
    'half', 'demi', 'extra', 'gorda', 'short', 'long', 'big', 'small',
    # common abbreviations
    'rob', 'chur', 'torp', 'gor',
    # numeric ring/length that can appear at tail
    '48', '50', '52', '54', '56', '58', '60', '64',
})

# Blend fields that are the same across all sizes of a line
_BLEND_FIELDS = (
    "Wrapper", "Binder", "Filler", "Country of Origin", "Factory",
    "Strength", "Flavor Notes",
    "Top Rating", "Rating Source", "Rating Source URL",
)

# Size-specific fields that must be researched fresh for each SKU
_SIZE_FIELDS = ("Shape", "Ring Gauge", "Length (in)", "MSRP", "MAP")


def _line_key(brand: str, description: str) -> str:
    """
    Return a normalised 'line identifier' by stripping trailing vitola/size words.

    Examples
    --------
    ("Dunbarton Tobacco & Trust", "Sobremesa Corona Grande")  → "dunbarton tobacco & trust|sobremesa"
    ("Dunbarton Tobacco & Trust", "Sobremesa Robusto Largo")  → "dunbarton tobacco & trust|sobremesa"
    ("Perdomo",                   "Perdomo Champagne Churchill") → "perdomo|perdomo champagne"
    ("Perdomo",                   "Perdomo Champagne Corona")    → "perdomo|perdomo champagne"
    ("Perdomo",                   "Perdomo BBA Mad. Churchill")  → "perdomo|perdomo bba mad."  ← different from Conn.
    """
    words = description.strip().split()
    # Strip from the right but keep at least the first word
    while len(words) > 1 and words[-1].lower().strip('.,()') in _VITOLA_WORDS:
        words.pop()
    line = ' '.join(words).strip()
    return f"{str(brand).strip().lower()}|{line.lower()}"


def _build_line_cache(cache: dict[str, dict]) -> dict[str, dict]:
    """Index existing cache records by line key → first matching record."""
    line_cache: dict[str, dict] = {}
    for record in cache.values():
        lk = _line_key(record.get("Brand", ""), record.get("Description", ""))
        if lk not in line_cache:
            line_cache[lk] = record
    return line_cache


# ── XLSX schema ──────────────────────────────────────────────────────────────
COLUMNS = [
    "Item Number",
    "UPC",
    "Description",
    "Brand",
    "Parent Company",
    # blend
    "Wrapper",
    "Binder",
    "Filler",
    "Country of Origin",
    "Factory",
    # character
    "Strength",       # Mild / Medium-Full / Full etc.
    "Shape",          # Robusto, Toro, Churchill …
    "Ring Gauge",
    "Length (in)",
    # marketing
    "Flavor Notes",        # synthesized tasting notes
    # ratings
    "Top Rating",          # highest rating found (e.g. "93" or "Buy a Box")
    "Rating Source",       # site name, e.g. "Cigar Aficionado"
    "Rating Source URL",   # direct URL to the review/rating page
    "MSRP",
    "MAP",
    # meta
    "Release Date",        # when the cigar shipped to retailers, e.g. "March 2024"
    "Source URLs",
    "Research Date",
    "Notes",
]

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")  # dark navy
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")   # light blue stripe
BODY_FONT    = Font(name="Calibri", size=10)

# ── agent system prompts ─────────────────────────────────────────────────────
# Used when blend info is already known — only look up size specs and MSRP.
SIZE_ONLY_PROMPT = """You are a cigar research specialist.
The blend details for this cigar line are already known. Your ONLY job is to find
the vitola specs (shape name, ring gauge, length) and single-stick MSRP/MAP for
this specific size. Use web_search with queries like:
  "[brand] [full product name] ring gauge length price"
  "site:cigars.com [full product name]"

Return ONLY a JSON object with these keys (null if not found):
{
  "shape":         string or null,
  "ring_gauge":    number or null,
  "length_inches": number or null,
  "msrp":          number or null,
  "map_price":     number or null,
  "source_urls":   [string]
}
Return ONLY the JSON — no markdown, no explanation."""

SYSTEM_PROMPT = """You are a cigar research specialist working for a premium cigar retailer.
Your job is to look up accurate, detailed information about a specific cigar SKU.

Use the web_search tool to search for information. Good queries to try:
  1. "[brand] [product name] wrapper binder filler"
  2. "site:halfwheel.com [brand] [product name]"
  3. "[brand] [product name] MSRP tasting notes"
  4. "[brand] [product name] official cigar details"

Extract and return a JSON object with EXACTLY these keys (use null for unknown):
{
  "wrapper":           string or null,
  "binder":            string or null,
  "filler":            string or null,
  "country_of_origin": string or null,   // country where rolled/made
  "factory":           string or null,
  "strength":          string or null,   // e.g. "Medium", "Full", "Medium-Full"
  "shape":             string or null,   // e.g. "Robusto", "Toro", "Churchill"
  "ring_gauge":        number or null,
  "length_inches":     number or null,
  "flavor_notes":      string or null,   // 1-3 sentence summary synthesized from manufacturer copy and reviewer notes; lead with the dominant profile, then secondary notes, then finish
  "top_rating":        string or null,   // highest rating found across all sources, as-is (e.g. "93", "94/100", "Buy a Box", "4.5/5")
  "rating_source":     string or null,   // name of the site that gave the top rating (e.g. "Cigar Aficionado", "Halfwheel", "Smoke Inn")
  "rating_source_url": string or null,   // direct URL to that review or rating page
  "msrp":              number or null,   // single stick retail price USD
  "map_price":         number or null,   // minimum advertised price USD (if known)
  "release_date":      string or null,   // when cigar shipped to retailers, e.g. "March 2024", "Spring 2022", "2019"
  "source_urls":       [string],         // URLs used as sources
  "notes":             string or null    // any caveats or extra info
}

IMPORTANT:
- HALLUCINATION RULE: Every field must be grounded in text you found in a web search result during THIS session. If a value cannot be confirmed by an actual search result, output null — never estimate, guess, or draw on model training knowledge. This applies especially to MSRP, blend components (wrapper/binder/filler), ratings, ring gauge, and length. If you find yourself writing a value without being able to cite a specific search result that contained it, use null instead.
- flavor_notes: synthesize ONLY from text found in actual search results (manufacturer copy, retailer descriptions, reviewer notes). Do not write flavor notes from model memory. Lead with the dominant flavor profile, follow with secondary notes, end with the finish. Use plain language a retail customer would understand — avoid marketing filler like "complex" or "unique experience". If no flavor descriptions were found in search results, use null. Example: "Rich dark chocolate and espresso up front with underlying earth and leather. Secondary notes of dried fruit and cedar. Long, peppery finish."
- top_rating / rating_source / rating_source_url: search for "[brand] [product] review rating score" and look across Cigar Aficionado, Halfwheel, Cigar Journal, Cigar Insider, Smoke Inn, Famous Smoke Shop, and similar sites. Pick the SINGLE highest rating found. For numeric scales (1-100 or 1-5 stars), prefer the highest number. For qualitative scales ("Buy a Box" > "Buy One" > "Pass" > "Don't Buy"), prefer the most positive. Report the rating exactly as the source states it — do not normalize or convert. If no rating is found, use null for all three fields.
- release_date: the month/year the cigar first shipped to retailers. Check review dates, press releases, and IPCPR/PCA trade show announcements. Use the format "Month YYYY" (e.g. "March 2024"), "Season YYYY" (e.g. "Spring 2022"), or just "YYYY" if only the year is known. Estimates are acceptable — note them in the notes field. Use null only if completely unknown.
- msrp is the single-stick retail price, not box price. Must come from a search result showing an actual price — never estimate based on brand tier.
- If multiple sizes exist, focus on the size matching the product name
- Always include at least 1 source URL when web search is used
- Return ONLY the JSON object, no markdown fences, no extra text
"""

# ── web search tool (Anthropic native — no extra API key needed) ──────────────
# Anthropic executes searches server-side; we just declare the tool and loop
# until end_turn.  Billed at $10/1,000 searches via your Anthropic account.
TOOL_DEF = {
    "type": "web_search_20250305",
    "name": "web_search",
    "max_uses": 5,   # cap per research call to control cost
}


# ── Rate-limit callback (per-thread) ─────────────────────────────────────────
# Allows the UI layer to be notified when a rate-limit wait starts without
# threading the callback through every call site.  Set once per worker thread
# via set_rate_limit_cb() before any tool work begins.

_tl = threading.local()


def set_rate_limit_cb(fn) -> None:
    """Register a callable(wait_secs, attempt) for the current thread."""
    _tl.rate_limit_cb = fn


def _fire_rate_limit_cb(wait: int, attempt: int) -> None:
    fn = getattr(_tl, "rate_limit_cb", None)
    if fn:
        try:
            fn(wait, attempt)
        except Exception:
            pass


# ── Rate-limit-aware API call ─────────────────────────────────────────────────

def _create_with_backoff(client: anthropic.Anthropic, **kwargs) -> anthropic.types.Message:
    """
    Call client.messages.create with exponential backoff on rate-limit (429) errors.
    Waits 60 s on the first hit, doubling each retry up to 4 attempts.
    Fires the per-thread rate-limit callback (if set) before each sleep so the
    UI can warn the user.
    """
    wait = 60
    for attempt in range(4):
        try:
            return client.messages.create(**kwargs)
        except anthropic.RateLimitError as exc:
            if attempt == 3:
                raise
            logging.warning(
                "Rate limit hit — waiting %d s before retry %d/3  (%s)",
                wait, attempt + 1, exc,
            )
            _fire_rate_limit_cb(wait, attempt + 1)
            print(f"  ⏳ Rate limit — waiting {wait}s…", flush=True)
            time.sleep(wait)
            wait = min(wait * 2, 300)  # cap at 5 min


# ── JSON extraction helper ────────────────────────────────────────────────────

import re as _re

def _extract_json(text: str) -> dict:
    """
    Robustly pull a JSON object out of a model response that may contain:
      • plain JSON with no fences
      • ```json ... ``` or ``` ... ``` fences
      • prose before/after the JSON block
      • a truncated response (token limit hit mid-JSON — no closing brace)

    Returns a (possibly empty) dict.
    """
    if not text:
        return {}

    text = text.strip()

    # Strategy 1 – find a complete {...} block anywhere in the text
    brace_match = _re.search(r'\{[\s\S]*\}', text)
    if brace_match:
        candidate = brace_match.group(0)
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            pass
        # Complete block found but still invalid — try repair
        try:
            return _repair_truncated_json(candidate)
        except Exception:
            pass

    # Strategy 2 – truncated: find opening brace, repair everything from there
    first_brace = text.find('{')
    if first_brace != -1:
        try:
            return _repair_truncated_json(text[first_brace:])
        except Exception:
            pass

    # Strategy 3 – strip ``` fences and try the inner content directly
    fenced = _re.sub(r'^```(?:json)?\s*', '', text, flags=_re.MULTILINE)
    fenced = _re.sub(r'```\s*$', '', fenced, flags=_re.MULTILINE).strip()
    try:
        return json.loads(fenced)
    except json.JSONDecodeError:
        pass

    logging.warning(
        "Could not parse JSON from agent response. Raw text (first 300 chars): %s",
        text[:300].replace('\n', ' ')
    )
    return {}


def _repair_truncated_json(text: str) -> dict:
    """
    Best-effort recovery for JSON cut off by a token limit.
    Closes any unclosed string literals, then balances braces/brackets.
    """
    import re

    # Drop the last incomplete key-value pair (everything after the last full comma)
    last_comma = text.rfind(',')
    if last_comma != -1:
        text = text[:last_comma]

    # Count open braces and brackets
    depth_curly  = text.count('{') - text.count('}')
    depth_square = text.count('[') - text.count(']')

    # Close any open string (odd number of unescaped quotes after last colon)
    if text.rstrip()[-1:] not in ('"', '}', ']', '0123456789'):
        text = text.rstrip().rstrip(',') + '"'

    text += ']' * max(depth_square, 0)
    text += '}' * max(depth_curly,  0)

    return json.loads(text)


# ── core researcher ───────────────────────────────────────────────────────────
class CigarResearcher:
    """Researches individual cigar SKUs and caches results in XLSX."""

    def __init__(self, model: str = "claude-sonnet-4-6") -> None:
        self.client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        self.model  = model
        self._ensure_research_file()

    # ── XLSX helpers ─────────────────────────────────────────────────────────

    def _ensure_research_file(self) -> None:
        """Create the research XLSX with headers if it doesn't exist, or migrate if columns changed."""
        if RESEARCH_FILE.exists():
            self._migrate_columns()
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cigar Research"

        # Header row
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(1, col_idx, col_name)
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Column widths
        widths = {
            "Item Number": 14, "UPC": 16, "Description": 38, "Brand": 22,
            "Parent Company": 28,
            "Wrapper": 24, "Binder": 22, "Filler": 24,
            "Country of Origin": 20, "Factory": 24,
            "Strength": 14, "Shape": 14, "Ring Gauge": 12, "Length (in)": 12,
            "Flavor Notes": 60,
            "Top Rating": 14, "Rating Source": 22, "Rating Source URL": 50,
            "MSRP": 10, "MAP": 10,
            "Release Date": 18,
            "Source URLs": 50, "Research Date": 16, "Notes": 40,
        }
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        wb.save(RESEARCH_FILE)
        logging.info("Created %s", RESEARCH_FILE)

    def load_cache(self) -> dict[str, dict]:
        """Return cached research keyed by (Description, Brand) tuple-string."""
        wb = openpyxl.load_workbook(RESEARCH_FILE, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {}
        cache: dict[str, dict] = {}
        for row in rows[1:]:
            record = dict(zip(COLUMNS, row))
            key = _cache_key(record.get("Description", ""), record.get("Brand", ""))
            cache[key] = record
        return cache

    def _migrate_columns(self) -> None:
        """Add any new columns to an existing research file without losing data."""
        wb = openpyxl.load_workbook(RESEARCH_FILE)
        ws = wb.active
        existing = [cell.value for cell in ws[1]]
        missing = [c for c in COLUMNS if c not in existing]
        if not missing:
            return

        logging.info("Migrating Cigar_Research.xlsx: adding columns %s", missing)
        widths = {
            "Top Rating": 14, "Rating Source": 22, "Rating Source URL": 50,
        }
        for col_name in missing:
            col_idx = COLUMNS.index(col_name) + 1
            ws.insert_cols(col_idx)
            cell = ws.cell(1, col_idx, col_name)
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)

        wb.save(RESEARCH_FILE)
        logging.info("Migration complete.")

    def save_result(self, result: dict) -> None:
        """Append or update one result row in the research XLSX."""
        wb = openpyxl.load_workbook(RESEARCH_FILE)
        ws = wb.active

        # Check if this key already exists → update in-place
        key = _cache_key(result.get("Description", ""), result.get("Brand", ""))
        target_row = None
        for row in ws.iter_rows(min_row=2):
            desc  = str(row[COLUMNS.index("Description")].value or "")
            brand = str(row[COLUMNS.index("Brand")].value or "")
            if _cache_key(desc, brand) == key:
                target_row = row[0].row
                break

        if target_row is None:
            target_row = ws.max_row + 1

        # Write values
        is_alt = (target_row % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(target_row, col_idx, result.get(col_name))
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("Flavor Notes", "Source URLs", "Notes")))
            if is_alt:
                cell.fill = ALT_FILL

        wb.save(RESEARCH_FILE)

    # ── agent loop ────────────────────────────────────────────────────────────

    def _research_size_only(
        self,
        description: str,
        brand: str,
        item_number: str | None,
        parent_company: str | None,
        sibling: dict,
        upc: str = "",
        search_id: str = "",
    ) -> dict[str, Any]:
        """
        Fast path: blend already known from a cached sibling.
        Only searches for shape, ring gauge, length, and MSRP for this size.
        """
        id_label = f"UPC {search_id}" if search_id else (f"Item# {item_number}" if item_number else "N/A")
        user_message = (
            f"Find vitola specs and single-stick MSRP for:\n"
            f"  Description : {description}\n"
            f"  Brand       : {brand}\n"
            f"  {id_label}\n\n"
            f"Known blend (same line, different size):\n"
            f"  Wrapper: {sibling.get('Wrapper')}\n"
            f"  Binder:  {sibling.get('Binder')}\n"
            f"  Filler:  {sibling.get('Filler')}\n"
        )

        messages = [{"role": "user", "content": user_message}]
        raw_json: dict = {}

        for _turn in range(4):  # shorter loop — just size + price
            response = _create_with_backoff(
                self.client,
                model=self.model,
                max_tokens=512,
                system=SIZE_ONLY_PROMPT,
                tools=[TOOL_DEF],
                messages=messages,
            )
            messages.append({"role": "assistant", "content": response.content})

            if response.stop_reason == "end_turn":
                for block in reversed(response.content):
                    if hasattr(block, "text"):
                        raw_json = _extract_json(block.text)
                        break
                break

            if response.stop_reason in ("tool_use", "max_tokens"):
                # API requires user/assistant alternation. Add a continuation
                # user turn to avoid "assistant prefill" error on next call.
                messages.append({"role": "user", "content": "Continue and provide the final JSON result."})
                continue

            break  # unexpected stop reason

        source_urls = raw_json.get("source_urls", [])
        if isinstance(source_urls, list):
            source_urls = "\n".join(source_urls)

        return {
            "Item Number":       item_number,
            "UPC":               upc or None,
            "Description":       description,
            "Brand":             brand,
            "Parent Company":    parent_company,
            # inherited blend fields
            "Wrapper":           sibling.get("Wrapper"),
            "Binder":            sibling.get("Binder"),
            "Filler":            sibling.get("Filler"),
            "Country of Origin": sibling.get("Country of Origin"),
            "Factory":           sibling.get("Factory"),
            "Strength":          sibling.get("Strength"),
            "Flavor Notes":      sibling.get("Flavor Notes"),
            "Top Rating":        sibling.get("Top Rating"),
            "Rating Source":     sibling.get("Rating Source"),
            "Rating Source URL": sibling.get("Rating Source URL"),
            # size-specific fields
            "Shape":             raw_json.get("shape"),
            "Ring Gauge":        raw_json.get("ring_gauge"),
            "Length (in)":       raw_json.get("length_inches"),
            "MSRP":              raw_json.get("msrp"),
            "MAP":               raw_json.get("map_price"),
            "Release Date":      sibling.get("Release Date"),
            "Source URLs":       source_urls,
            "Research Date":     str(date.today()),
            "Notes":             f"Blend inherited from: {sibling.get('Description')}",
        }

    def research_cigar(
        self,
        description: str,
        brand: str,
        item_number: str | None = None,
        parent_company: str | None = None,
        upc: str = "",
        search_id: str = "",
    ) -> dict[str, Any]:
        """
        Research a single cigar SKU.  Returns a flat dict matching COLUMNS.

        upc: barcode from inventory (stored in research file).
        search_id: UPC if available, otherwise Item Number — used in web queries.

        If a sibling from the same blend line is already cached, inherit all
        blend fields (wrapper, binder, flavor notes, ratings, etc.) and only
        do a short web search for the size-specific specs (shape, ring gauge,
        length, MSRP).  Saves roughly 75 % of API calls on multi-size lines.
        """
        logging.info("Researching: %s (%s) [upc=%s]", description, brand, upc or "—")

        # ── Line-cache check ──────────────────────────────────────────────────
        lkey       = _line_key(brand, description)
        line_cache = _build_line_cache(self.load_cache())
        sibling    = line_cache.get(lkey)

        if sibling:
            logging.info(
                "  Line match found: inheriting blend from '%s'",
                sibling.get("Description"),
            )
            result = self._research_size_only(
                description, brand, item_number, parent_company, sibling, upc, search_id
            )
            self.save_result(result)
            return result

        # ── Full research ─────────────────────────────────────────────────────
        # UPC is the most reliable search identifier; fall back to item number
        id_label = f"UPC {search_id}" if search_id else (f"Item# {item_number}" if item_number else "N/A")
        user_message = (
            f"Please research this cigar SKU:\n"
            f"  Description : {description}\n"
            f"  Brand       : {brand}\n"
            f"  {id_label}\n\n"
            f"The UPC or item number above can be used as a search term for precise results. "
            f"Find wrapper, binder, filler, country of origin, factory, strength, "
            f"shape/vitola, ring gauge, length, manufacturer flavor notes, and "
            f"single-stick MSRP/MAP. Use web search for accurate current info."
        )

        messages = [{"role": "user", "content": user_message}]
        raw_json: dict = {}

        for _turn in range(8):
            response = _create_with_backoff(
                self.client,
                model=self.model,
                max_tokens=2048,
                system=SYSTEM_PROMPT,
                tools=[TOOL_DEF],
                messages=messages,
            )
            messages.append({"role": "assistant", "content": response.content})

            if response.stop_reason == "end_turn":
                for block in reversed(response.content):
                    if hasattr(block, "text"):
                        raw_json = _extract_json(block.text)
                        break
                break

            # Native web_search tool: Anthropic executes the search server-side.
            # The tool_result is already included in response.content — we just
            # need to keep looping until end_turn.
            # The API requires user/assistant alternation — add a continuation
            # user turn to avoid "assistant prefill" error on the next call.
            if response.stop_reason in ("tool_use", "max_tokens"):
                messages.append({"role": "user", "content": "Continue and provide the final JSON result."})
                continue

            break  # unexpected stop reason

        # Map agent keys → COLUMN names
        source_urls = raw_json.get("source_urls", [])
        if isinstance(source_urls, list):
            source_urls = "\n".join(source_urls)

        result = {
            "Item Number":       item_number,
            "UPC":               upc or None,
            "Description":       description,
            "Brand":             brand,
            "Parent Company":    parent_company,
            "Wrapper":           raw_json.get("wrapper"),
            "Binder":            raw_json.get("binder"),
            "Filler":            raw_json.get("filler"),
            "Country of Origin": raw_json.get("country_of_origin"),
            "Factory":           raw_json.get("factory"),
            "Strength":          raw_json.get("strength"),
            "Shape":             raw_json.get("shape"),
            "Ring Gauge":        raw_json.get("ring_gauge"),
            "Length (in)":       raw_json.get("length_inches"),
            "Flavor Notes":      raw_json.get("flavor_notes"),
            "Top Rating":        raw_json.get("top_rating"),
            "Rating Source":     raw_json.get("rating_source"),
            "Rating Source URL": raw_json.get("rating_source_url"),
            "MSRP":              raw_json.get("msrp"),
            "MAP":               raw_json.get("map_price"),
            "Release Date":      raw_json.get("release_date"),
            "Source URLs":       source_urls,
            "Research Date":     str(date.today()),
            "Notes":             raw_json.get("notes"),
        }

        self.save_result(result)
        return result

    # ── batch helper ──────────────────────────────────────────────────────────

    def batch_research(
        self,
        limit: int | None = None,
        force: bool = False,
        rate_limit_s: float = 15.0,
        category: str = "Cigars",
        since: str | None = None,
        sort_by_sales: bool = False,
        exclude_brands: tuple[str, ...] = ("Smoke Shoppe",),
        exclude_parent_companies: tuple[str, ...] = (),
    ) -> list[dict]:
        """
        Research inventory cigars, optionally filtered and sorted by sales.

        since         – timeframe string: "2025", "last 6 months", "last 90 days",
                        "Q2 2024", or "2024-01 to 2024-06".  Filters to items
                        that had at least one sale within that window.
        sort_by_sales – if True, research highest-selling items first (units sold
                        in the `since` window, or all-time if since is None).
        exclude_brands          – brand names to skip (default: Smoke Shoppe house brands).
        exclude_parent_companies – parent companies to skip (default: Smoke Shoppe).
                                   Catches any sub-brand under the house parent.
        """
        from tools.inventory_tool import run_inventory_sql_df
        conditions = []
        if category:
            conditions.append(f"Category = '{category}'")
        if exclude_brands:
            brands_sql = ", ".join(f"'{b.replace(chr(39), chr(39)*2)}'" for b in exclude_brands)
            conditions.append(f"Brand NOT IN ({brands_sql})")
        if exclude_parent_companies:
            pcs_sql = ", ".join(f"'{p.replace(chr(39), chr(39)*2)}'" for p in exclude_parent_companies)
            conditions.append(f'"Parent Company" NOT IN ({pcs_sql})')
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        inv = run_inventory_sql_df(
            f'SELECT "Item Number", Description, Brand, "Parent Company", UPC '
            f"FROM inventory {where}",
            file_path=str(INVENTORY_FILE),
        )

        # ── Sales filter / sort ───────────────────────────────────────────────
        if since or sort_by_sales:
            inv = self._apply_sales_filter(inv, since=since, sort_by_sales=sort_by_sales)

        cache = {} if force else self.load_cache()
        results = []
        skipped = 0
        count = 0

        for _, row in inv.iterrows():
            description    = str(row.get("Description", "") or "")
            brand          = str(row.get("Brand", "") or "")
            item_number    = str(row.get("Item Number", "") or "")
            parent_company = str(row.get("Parent Company", "") or "")
            upc            = str(row.get("UPC", "") or "").strip()
            if upc.endswith(".0"):
                upc = upc[:-2]
            search_id = upc if upc and upc != "nan" else item_number

            if not description or description == "nan":
                continue

            key = _cache_key(description, brand)
            if key in cache:
                skipped += 1
                continue

            units   = row.get("_units_sold", "")
            revenue = row.get("_revenue", "")
            if units != "" and revenue != "":
                units_str = f"  {int(units)} units / ${revenue:,.0f} rev"
            elif units != "":
                units_str = f"  {int(units)} units sold"
            else:
                units_str = ""
            try:
                result = self.research_cigar(description, brand, item_number, parent_company, upc=upc, search_id=search_id)
                results.append(result)
                count += 1
                print(f"[{count}] ✓ {description[:50]}{units_str}", flush=True)
            except Exception as exc:
                logging.error("Error researching %s: %s", description, exc)
                print(f"[{count}] ✗ {description[:50]} — {exc}", flush=True)

            if limit and count >= limit:
                break

            time.sleep(rate_limit_s)

        print(f"\nDone. Researched: {count}  |  Already cached (skipped): {skipped}")
        return results

    @staticmethod
    def _apply_sales_filter(
        inv: "pd.DataFrame",
        since: str | None,
        sort_by_sales: bool,
    ) -> "pd.DataFrame":
        """Thin wrapper — see module-level _apply_sales_filter for full docs."""
        return _apply_sales_filter(inv, since, sort_by_sales)


# ── public lookup (used by sales agent & MCP server) ─────────────────────────

_researcher: CigarResearcher | None = None


def get_researcher() -> CigarResearcher:
    global _researcher
    if _researcher is None:
        _researcher = CigarResearcher()
    return _researcher


def lookup_cigar(
    description: str,
    brand: str = "",
    item_number: str = "",
    parent_company: str = "",
    upc: str = "",
    use_cache: bool = True,
) -> dict[str, Any]:
    """
    Main entry point for external callers (sales agent, MCP server).

    Returns a dict with all researched fields.  Hits the XLSX cache first;
    only calls the agent if the item hasn't been researched yet (or use_cache=False).
    UPC is preferred over item_number for web searches when available.
    """
    researcher = get_researcher()

    if use_cache:
        cache = researcher.load_cache()
        key = _cache_key(description, brand)
        if key in cache:
            return cache[key]

    clean_upc = upc.strip() if upc else ""
    search_id = clean_upc or item_number
    return researcher.research_cigar(description, brand, item_number, parent_company, upc=clean_upc, search_id=search_id)


def get_all_research() -> list[dict]:
    """Return every row in the research cache as a list of dicts."""
    researcher = get_researcher()
    cache = researcher.load_cache()
    return list(cache.values())


def research_status() -> dict:
    """Return a summary: how many items are cached vs. total inventory."""
    from tools.inventory_tool import run_inventory_sql
    total = run_inventory_sql(
        "SELECT COUNT(*) FROM inventory WHERE Category = 'Cigars'",
        file_path=str(INVENTORY_FILE),
    )[0][0]
    cache  = get_researcher().load_cache()
    return {
        "total_cigar_skus": total,
        "researched":       len(cache),
        "remaining":        max(0, total - len(cache)),
        "research_file":    str(RESEARCH_FILE),
    }


# ── util ──────────────────────────────────────────────────────────────────────

def _parse_since(since: str) -> tuple["pd.Timestamp", "pd.Timestamp"]:
    """
    Parse a human-readable timeframe string into (start, end) Timestamps.

    Supported formats
    -----------------
    "2025"              → full calendar year 2025
    "2024-06"           → June 2024 only
    "last 6 months"     → past 6 calendar months
    "last 90 days"      → past 90 days
    "last 2 years"      → past 2 years
    "Q1 2025"           → Jan–Mar 2025  (Q2=Apr-Jun, Q3=Jul-Sep, Q4=Oct-Dec)
    "2024-01 to 2024-06"→ explicit range (both YYYY-MM)
    """
    import re
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta

    s = since.strip()
    today = pd.Timestamp.today().normalize()

    # "last N days / weeks / months / years"
    m = re.fullmatch(r"last\s+(\d+)\s+(day|days|week|weeks|month|months|year|years)", s, re.I)
    if m:
        n, unit = int(m.group(1)), m.group(2).lower().rstrip("s")
        if unit == "day":
            start = today - timedelta(days=n)
        elif unit == "week":
            start = today - timedelta(weeks=n)
        elif unit == "month":
            start = today - relativedelta(months=n)
        else:
            start = today - relativedelta(years=n)
        return start, today

    # "Q1 2025" … "Q4 2025"
    m = re.fullmatch(r"Q([1-4])\s+(\d{4})", s, re.I)
    if m:
        q, yr = int(m.group(1)), int(m.group(2))
        start_month = (q - 1) * 3 + 1
        start = pd.Timestamp(yr, start_month, 1)
        end   = start + pd.offsets.QuarterEnd(0)
        return start, end

    # "YYYY-MM to YYYY-MM"
    m = re.fullmatch(r"(\d{4}-\d{2})\s+to\s+(\d{4}-\d{2})", s, re.I)
    if m:
        start = pd.Timestamp(m.group(1) + "-01")
        end   = pd.Timestamp(m.group(2) + "-01") + pd.offsets.MonthEnd(0)
        return start, end

    # "YYYY-MM"
    m = re.fullmatch(r"(\d{4})-(\d{2})", s)
    if m:
        start = pd.Timestamp(int(m.group(1)), int(m.group(2)), 1)
        end   = start + pd.offsets.MonthEnd(0)
        return start, end

    # "YYYY"
    m = re.fullmatch(r"(\d{4})", s)
    if m:
        yr = int(m.group(1))
        return pd.Timestamp(yr, 1, 1), pd.Timestamp(yr, 12, 31)

    raise ValueError(
        f"Unrecognised timeframe: {since!r}\n"
        "Supported: 'YYYY', 'YYYY-MM', 'last N months', 'last N days', "
        "'Q1 YYYY', 'YYYY-MM to YYYY-MM'"
    )


def _apply_sales_filter(
    inv: "pd.DataFrame",
    since: str | None,
    sort_by_sales: bool,
) -> "pd.DataFrame":
    """
    Join inventory with transactions to get units sold and revenue,
    optionally restricted to a date window, and sort by balanced
    qty+revenue score (descending) if sort_by_sales is True.

    Balanced sort: equal weight between a cigar's share of total units
    sold and its share of total revenue.  This prevents cheap high-volume
    items from dominating over more valuable lines.

    When ``since`` is set, items with zero sales in that window are
    excluded (inner join).  When only ``sort_by_sales`` is True, unsold
    items appear at the bottom (left join, score 0).

    Adds columns: _units_sold, _revenue, _balanced_score (if sort_by_sales).
    """
    import duckdb
    # Register the passed-in inventory slice and query transactions via DuckDB.
    # Only the three columns we need are pulled from transactions.
    conn = duckdb.connect()
    conn.register("inv", inv)

    date_filter = ""
    if since:
        start, end = _parse_since(since)
        date_filter = f"AND CAST(t.Date AS VARCHAR) >= '{start}' AND CAST(t.Date AS VARCHAR) <= '{end}'"

    tx_path = str(DATA_DIR / "Smoke_Shoppe_Transactions.xlsx")
    # Load only the columns we need from transactions
    import pandas as _pd
    txn = _pd.read_excel(tx_path, header=0,
                         usecols=["Date", "Item Number", "Quantity", "Item Amount"])
    txn["Date"] = _pd.to_datetime(txn["Date"], format="%m/%d/%y", errors="coerce")
    conn.register("transactions", txn)

    join_type = "INNER" if since else "LEFT"
    date_clause = ""
    if since:
        start, end = _parse_since(since)
        date_clause = f"AND t.Date >= '{start}' AND t.Date <= '{end}'"

    result = conn.execute(f"""
        SELECT
            inv.*,
            COALESCE(SUM(t.Quantity), 0)      AS _units_sold,
            COALESCE(SUM(t."Item Amount"), 0) AS _revenue
        FROM inv
        {join_type} JOIN transactions t
            ON CAST(inv."Item Number" AS VARCHAR) = CAST(t."Item Number" AS VARCHAR)
            {date_clause}
        GROUP BY ALL
        ORDER BY inv.rowid
    """).fetchdf()

    if result.empty and since:
        print(f"  ⚠ No transactions found for '{since}' — proceeding without filter.")

    if sort_by_sales:
        total_qty = result["_units_sold"].sum()
        total_rev = result["_revenue"].sum()
        qty_share = result["_units_sold"] / total_qty if total_qty > 0 else 0.0
        rev_share = result["_revenue"]    / total_rev if total_rev > 0 else 0.0
        result["_balanced_score"] = (qty_share + rev_share) / 2
        result = result.sort_values("_balanced_score", ascending=False)

    return result.reset_index(drop=True)


def _cache_key(description: str, brand: str) -> str:
    return f"{str(description).strip().lower()}|{str(brand).strip().lower()}"


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    logging.basicConfig(level=logging.WARNING, format="%(levelname)s: %(message)s")

    parser = argparse.ArgumentParser(description="Cigar Research Agent")
    parser.add_argument("description", nargs="?", help="Cigar description to research")
    parser.add_argument("brand",       nargs="?", default="", help="Cigar brand")
    parser.add_argument("--batch",       action="store_true", help="Research all uncached inventory cigars")
    parser.add_argument("--limit",       type=int,   default=None,  help="Max items in batch mode")
    parser.add_argument("--force",       action="store_true",       help="Re-research even if cached")
    parser.add_argument("--delay",       type=float, default=15.0,  help="Seconds between requests (default 15)")
    parser.add_argument("--since",       type=str,   default=None,
                        help="Only research cigars sold within this window. "
                             "Examples: '2025', 'last 6 months', 'last 90 days', 'Q2 2024', '2024-01 to 2024-06'")
    parser.add_argument("--top",         action="store_true",
                        help="Sort by balanced qty+revenue score (descending) within --since window, or all-time if omitted")
    parser.add_argument("--status",      action="store_true",       help="Show cache status and exit")
    parser.add_argument("--test-search", action="store_true",       help="Run a quick live search test and exit")
    args = parser.parse_args()

    if args.test_search:
        print("Testing Anthropic native web search…")
        result = lookup_cigar("Perdomo BBA Mad. Churchill", "Perdomo", use_cache=False)
        if result.get("Wrapper"):
            print(f"✓ Web search working.")
            print(f"  Wrapper : {result.get('Wrapper')}")
            print(f"  Strength: {result.get('Strength')}")
            print(f"  MSRP    : {result.get('MSRP')}")
        else:
            print("✗ Search ran but returned no wrapper data — check your ANTHROPIC_API_KEY.")
        sys.exit(0)

    if args.status:
        s = research_status()
        print(f"Total cigar SKUs : {s['total_cigar_skus']}")
        print(f"Already cached   : {s['researched']}")
        print(f"Still remaining  : {s['remaining']}")
        print(f"Research file    : {s['research_file']}")
        sys.exit(0)

    if args.batch:
        get_researcher().batch_research(
            limit=args.limit,
            force=args.force,
            rate_limit_s=args.delay,
            since=args.since,
            sort_by_sales=args.top,
        )
    elif args.description:
        result = lookup_cigar(args.description, args.brand, use_cache=not args.force)
        print(json.dumps(result, indent=2, default=str))
    else:
        parser.print_help()
