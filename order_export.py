"""
order_export.py
---------------
Export ordering-agent results to XLSX (openpyxl) or PDF (reportlab).

Public API:
    from order_export import to_xlsx, to_pdf
    path = to_xlsx(result)                    # → exports/order_2026-06-10.xlsx
    path = to_pdf(result)                     # → exports/order_2026-06-10.pdf
    path = to_xlsx(result, "/tmp/order.xlsx") # explicit path

CLI usage (via ordering_agent.py --export):
    python main.py order --export xlsx
    python main.py order --export pdf
    python main.py order --export both
    python main.py order --export xlsx --export-path /tmp/order.xlsx
"""
from __future__ import annotations

from datetime import date
from pathlib import Path


# ── path helpers ──────────────────────────────────────────────────────────────

def _default_stem() -> str:
    return f"order_{date.today()}"


def _resolve_path(path: str | Path | None, suffix: str) -> Path:
    if path is None:
        exports = Path("exports")
        exports.mkdir(exist_ok=True)
        return exports / f"{_default_stem()}{suffix}"
    p = Path(path)
    if p.is_dir():
        return p / f"{_default_stem()}{suffix}"
    return p


# ── brand palette ─────────────────────────────────────────────────────────────

_BROWN  = "4A2C17"   # dark-brown header
_AMBER  = "C8860A"   # amber accent / sub-header
_CREAM  = "FFF8EE"   # alternate row fill
_WHITE  = "FFFFFF"
_GOLD   = "EDD9A3"   # total-row fill
_GREEN  = "27AE60"
_RED    = "C0392B"
_YELLOW = "FCF3CF"   # warning fill


# ── XLSX export ───────────────────────────────────────────────────────────────

def to_xlsx(result: dict, path: str | Path | None = None) -> Path:
    """Write result to a formatted Excel workbook.  Returns the path written."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: uv add openpyxl") from exc

    out = _resolve_path(path, ".xlsx")

    meta    = result.get("metadata", {})
    rec     = result.get("recommendation", {})
    restock = result.get("restock", {})
    grouped = result.get("order_by_parent_company", {})

    run_date     = date.today().strftime("%B %d, %Y")
    horizon_days = meta.get("horizon_days", 30)
    order_budget = meta.get("order_budget")
    new_bud      = meta.get("new_cigar_budget")
    rest_bud     = meta.get("restock_budget")
    combined     = meta.get("combined_order_cost")
    seasonal     = meta.get("seasonal_context", "")
    strategy     = rec.get("ordering_strategy", "")

    # ── style factory helpers ─────────────────────────────────────────────────

    def _fill(hex_: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_)

    def _font(bold=False, color="000000", size=10) -> Font:
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def _thin_border() -> Border:
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _banner(ws, row: int, text: str, span: int, height: int = 24) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.fill = _fill(_AMBER)
        c.font = _font(bold=True, color="FFFFFF", size=13)
        c.alignment = Alignment(horizontal="left", vertical="center")
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        ws.row_dimensions[row].height = height

    def _subhead(ws, row: int, text: str, span: int) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.fill = _fill(_BROWN)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        ws.row_dimensions[row].height = 18

    def _header_row(ws, row: int, headers: list[str], col_start: int = 1) -> None:
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.fill = _fill(_BROWN)
            c.font = _font(bold=True, color="FFFFFF", size=9)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _thin_border()
        ws.row_dimensions[row].height = 30

    def _data_row(ws, row: int, values: list, alt: bool = False) -> None:
        fill = _fill(_CREAM if alt else _WHITE)
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.font = _font(size=9)
            c.border = _thin_border()
            c.alignment = Alignment(vertical="top", wrap_text=True)

    def _total_row(ws, row: int, values: list) -> None:
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = _fill(_GOLD)
            c.font = _font(bold=True, size=9)
            c.border = _thin_border()
            c.alignment = Alignment(vertical="center")

    def _money(ws, row: int, col: int) -> None:
        ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 46

    _banner(ws, 1, "SMOKE SHOPPE — ORDER RECOMMENDATION", 2)
    ws.row_dimensions[1].height = 28

    rows_info: list[tuple[str, object]] = [
        ("Generated",        run_date),
        ("Planning horizon", f"{horizon_days} days"),
        ("Total budget",     f"${order_budget:,.0f}" if order_budget else "auto"),
        ("  → New cigars",   f"${new_bud:,.2f}"      if new_bud is not None else "—"),
        ("  → Restock",      f"${rest_bud:,.2f}"     if rest_bud is not None else "—"),
        ("Strategy",         strategy.upper() if strategy else "—"),
        ("Seasonal context", seasonal or "—"),
        ("", ""),
        ("Summary",          rec.get("summary", "")),
        ("Branch consensus", rec.get("branch_consensus", "")),
    ]

    r = 2
    for label, val in rows_info:
        ws.cell(row=r, column=1, value=label).font = _font(bold=bool(label), size=10)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if label in ("Summary", "Branch consensus"):
            ws.row_dimensions[r].height = 42
        r += 1

    for w in result.get("budget_warnings", []):
        c = ws.cell(row=r, column=1, value=w)
        c.fill = _fill(_YELLOW)
        c.font = _font(color="7D6608", size=9)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1

    r += 1
    _subhead(ws, r, "COST BREAKDOWN", 2)
    r += 1
    for label, amount in [
        ("New cigar recommendations",   rec.get("total_order_cost") or 0),
        ("Restock (ordered)",           restock.get("total_cost_ordered") or 0),
        ("Combined wholesale cost",     combined or 0),
    ]:
        ws.cell(row=r, column=1, value=label).font = _font(bold=(label == "Combined wholesale cost"), size=10)
        c = ws.cell(row=r, column=2, value=amount)
        c.font = _font(bold=(label == "Combined wholesale cost"), size=10)
        c.number_format = '"$"#,##0.00'
        r += 1

    ws.freeze_panes = "A2"

    # ── Sheet 2: New Cigars ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("New Cigars")
    hdrs2 = ["Rank", "Cigar", "Brand", "Vitola", "Conviction",
             "Branches", "Boxes", "Box Size", "MSRP/Stick",
             "Est. Wholesale", "Rationale", "Risk"]
    widths2 = [5, 26, 22, 12, 10, 28, 6, 8, 10, 13, 48, 28]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    _banner(ws2, 1, "NEW CIGAR RECOMMENDATIONS", len(hdrs2))
    _header_row(ws2, 2, hdrs2)
    ws2.freeze_panes = "A3"

    r = 3
    _urgency_conv = {"high": _GREEN, "medium": _AMBER, "bold": _RED}
    for idx, o in enumerate(rec.get("recommended_orders", [])):
        conviction = o.get("conviction", "")
        vals = [
            o.get("rank", idx + 1),
            o.get("name", ""),
            o.get("brand", ""),
            o.get("vitola", ""),
            conviction.upper(),
            ", ".join(o.get("branches_agreed", [])),
            o.get("boxes", ""),
            o.get("box_size", ""),
            o.get("msrp_per_stick"),
            o.get("cost_estimate"),
            o.get("rationale", ""),
            o.get("watch_out_for") or "",
        ]
        _data_row(ws2, r, vals, alt=bool(idx % 2))
        ws2.cell(row=r, column=5).font = Font(
            bold=True, color=_urgency_conv.get(conviction, "000000"), size=9, name="Calibri"
        )
        _money(ws2, r, 9)
        _money(ws2, r, 10)
        ws2.row_dimensions[r].height = 52
        r += 1

    if rec.get("recommended_orders"):
        total_new = rec.get("total_order_cost") or 0
        _total_row(ws2, r, ["", "", "", "", "", "", "", "", "TOTAL", total_new, "", ""])
        _money(ws2, r, 10)

    # ── Sheet 3: Restock ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Restock")
    hdrs3 = ["Urgency", "Description", "Brand", "Status", "On Hand",
             "Days Left", "Vel./mo", "Boxes", "Box Size", "Units",
             "Cost/Stick", "Wholesale", "Seasonal ×"]
    widths3 = [10, 30, 22, 14, 7, 7, 9, 6, 8, 6, 10, 13, 10]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    _banner(ws3, 1, "LOW-STOCK REORDER", len(hdrs3))
    _header_row(ws3, 2, hdrs3)
    ws3.freeze_panes = "A3"

    _urgency_label = {
        "critical": "🔴 Critical", "high": "🟠 High",
        "medium": "🟡 Medium",    "low":  "🟢 Low",
    }

    r = 3

    def _write_restock_rows(ws, items: list, row_start: int, alt_offset: int = 0) -> int:
        rr = row_start
        for idx, item in enumerate(items):
            urgency = item.get("urgency", "")
            days    = item.get("days_until_stockout")
            sf      = item.get("seasonal_factor")
            vals = [
                _urgency_label.get(urgency, urgency),
                item.get("description", ""),
                item.get("brand", ""),
                item.get("status", "").replace("_", " "),
                item.get("on_hand", 0),
                f"{int(days)}d" if days is not None else "OOS",
                round(item.get("monthly_velocity", 0), 1),
                item.get("reorder_boxes", 0),
                item.get("box_size", 20),
                item.get("reorder_qty", 0),
                item.get("cost", 0),
                item.get("reorder_cost", 0),
                f"×{sf:.2f}" if sf else "—",
            ]
            _data_row(ws, rr, vals, alt=bool((idx + alt_offset) % 2))
            _money(ws, rr, 11)
            _money(ws, rr, 12)
            ws.row_dimensions[rr].height = 28
            rr += 1
        return rr

    r = _write_restock_rows(ws3, restock.get("items", []), r)

    if restock.get("items"):
        total_rest = restock.get("total_cost_ordered", 0)
        blanks = [""] * len(hdrs3)
        blanks[0] = "TOTAL"
        blanks[11] = total_rest
        _total_row(ws3, r, blanks)
        _money(ws3, r, 12)
        r += 1

    skipped = restock.get("items_skipped", [])
    if skipped:
        r += 1
        _subhead(ws3, r, f"DEFERRED — BUDGET EXHAUSTED ({len(skipped)} items)", len(hdrs3))
        r += 1
        _header_row(ws3, r, hdrs3)
        r += 1
        r = _write_restock_rows(ws3, skipped, r)

    # ── Sheet 4: By Vendor ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("By Vendor")
    widths4 = [34, 8, 32, 14, 9, 9, 14]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    _banner(ws4, 1, "CONSOLIDATED ORDER BY VENDOR", len(widths4))
    r = 2

    for parent, group in grouped.items():
        group_total = group.get("group_total_cost", 0)
        for col in range(1, len(widths4) + 1):
            c = ws4.cell(row=r, column=col)
            c.fill = _fill(_BROWN)
            c.border = _thin_border()
        ws4.cell(row=r, column=1, value=parent).font = _font(bold=True, color="FFFFFF", size=11)
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(widths4) - 1)
        c7 = ws4.cell(row=r, column=len(widths4), value=group_total)
        c7.fill = _fill(_BROWN)
        c7.font = _font(bold=True, color="FFFFFF", size=11)
        c7.number_format = '"$"#,##0.00'
        c7.alignment = Alignment(horizontal="right", vertical="center")
        ws4.row_dimensions[r].height = 20
        r += 1

        all_items = [
            ("[RESTOCK] " + item.get("description", ""),
             item.get("reorder_boxes", 0),
             f"× {item.get('box_size', 20)} sticks/box = {item.get('reorder_qty', 0)} sticks",
             item.get("urgency", "").capitalize(),
             item.get("reorder_cost", 0))
            for item in group.get("restock", [])
        ] + [
            ("[NEW] " + order.get("name", "") + (f" — {order.get('vitola','')}" if order.get("vitola") else ""),
             order.get("boxes", 0),
             f"× {order.get('box_size', 20)} sticks/box",
             order.get("conviction", "").capitalize(),
             order.get("cost_estimate", 0))
            for order in group.get("new_cigars", [])
        ]

        for idx, (desc, boxes, qty_str, label, cost) in enumerate(all_items):
            _data_row(ws4, r, [desc, boxes, qty_str, label, "", "", cost], alt=bool(idx % 2))
            _money(ws4, r, 7)
            ws4.row_dimensions[r].height = 22
            r += 1

        r += 1  # blank separator

    _total_row(ws4, r, ["GRAND TOTAL", "", "", "", "", "", combined or 0])
    _money(ws4, r, 7)

    wb.save(out)
    return out


# ── PDF export ────────────────────────────────────────────────────────────────

def to_pdf(result: dict, path: str | Path | None = None) -> Path:
    """Write result to a formatted PDF purchase order.  Returns the path written."""
    try:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.units import inch
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
        )
        from reportlab.platypus import HRFlowable
    except ImportError as exc:
        raise RuntimeError("reportlab is required: uv add reportlab") from exc

    out = _resolve_path(path, ".pdf")

    meta    = result.get("metadata", {})
    rec     = result.get("recommendation", {})
    restock = result.get("restock", {})
    grouped = result.get("order_by_parent_company", {})

    run_date     = date.today().strftime("%B %d, %Y")
    horizon_days = meta.get("horizon_days", 30)
    order_budget = meta.get("order_budget")
    new_bud      = meta.get("new_cigar_budget")
    rest_bud     = meta.get("restock_budget")
    combined     = meta.get("combined_order_cost")
    seasonal     = meta.get("seasonal_context", "")
    strategy     = rec.get("ordering_strategy", "")

    BROWN  = HexColor(f"#{_BROWN}")
    AMBER  = HexColor(f"#{_AMBER}")
    CREAM  = HexColor(f"#{_CREAM}")
    GOLD   = HexColor(f"#{_GOLD}")
    GREEN  = HexColor(f"#{_GREEN}")
    RED    = HexColor(f"#{_RED}")
    YELLOW = HexColor(f"#{_YELLOW}")

    doc = SimpleDocTemplate(
        str(out),
        pagesize=LETTER,
        leftMargin=0.65 * inch,
        rightMargin=0.65 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.65 * inch,
    )
    page_width = LETTER[0] - 1.3 * inch

    styles = getSampleStyleSheet()

    def _style(name, base="Normal", **kwargs) -> ParagraphStyle:
        return ParagraphStyle(name, parent=styles[base], **kwargs)

    h1      = _style("H1",      fontSize=16, textColor=white,  leading=20, spaceAfter=0)
    h2      = _style("H2",      fontSize=11, textColor=white,  leading=14, spaceAfter=0)
    h3      = _style("H3",      fontSize=9,  textColor=BROWN,  leading=13, fontName="Helvetica-Bold")
    body    = _style("Body",    fontSize=8,  leading=11)
    small   = _style("Small",   fontSize=7,  leading=9,        textColor=HexColor("#555555"))
    wrap    = _style("Wrap",    fontSize=7,  leading=9,        wordWrap="LTR")
    bold_s  = _style("BoldS",   fontSize=8,  fontName="Helvetica-Bold", leading=10)
    right_s = _style("Right",   fontSize=8,  alignment=TA_RIGHT, leading=10)

    story = []

    # ── Title banner ──────────────────────────────────────────────────────────
    banner_table = Table(
        [[Paragraph("SMOKE SHOPPE — ORDER RECOMMENDATION", h1)]],
        colWidths=[page_width],
    )
    banner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), AMBER),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 6))

    # ── Run metadata ──────────────────────────────────────────────────────────
    meta_rows = [
        ("Generated",        run_date),
        ("Horizon",          f"{horizon_days} days"),
        ("Total budget",     f"${order_budget:,.0f}" if order_budget else "auto"),
        ("  New cigars",     f"${new_bud:,.2f}"      if new_bud is not None else "—"),
        ("  Restock",        f"${rest_bud:,.2f}"     if rest_bud is not None else "—"),
        ("Strategy",         strategy.upper() if strategy else "—"),
    ]
    if seasonal:
        meta_rows.append(("Seasonal context", seasonal))

    meta_data = [[Paragraph(f"<b>{k}</b>", body), Paragraph(v, body)] for k, v in meta_rows]
    meta_tbl = Table(meta_data, colWidths=[1.4 * inch, page_width - 1.4 * inch])
    meta_tbl.setStyle(TableStyle([
        ("FONTSIZE",     (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(meta_tbl)

    if rec.get("summary"):
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"<b>Summary:</b> {rec['summary']}", body))

    for w in result.get("budget_warnings", []):
        story.append(Spacer(1, 4))
        warn_tbl = Table([[Paragraph(f"⚠ {w}", small)]], colWidths=[page_width])
        warn_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), YELLOW),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ]))
        story.append(warn_tbl)

    story.append(Spacer(1, 10))

    # ── Section helper ────────────────────────────────────────────────────────
    def _section_header(title: str) -> Table:
        t = Table([[Paragraph(title, h2)]], colWidths=[page_width])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), BROWN),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        return t

    def _col_header_style() -> list:
        return [
            ("BACKGROUND",    (0, 0), (-1, 0), BROWN),
            ("TEXTCOLOR",     (0, 0), (-1, 0), white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 7),
            ("TOPPADDING",    (0, 0), (-1, 0), 4),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
            ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ]

    def _row_style(n_data_rows: int, alt_col=None) -> list:
        style = [
            ("FONTSIZE",      (0, 1), (-1, -1), 7),
            ("TOPPADDING",    (0, 1), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
            ("VALIGN",        (0, 1), (-1, -1), "TOP"),
            ("GRID",          (0, 0), (-1, -1), 0.3, HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, CREAM]),
        ]
        return style

    # ── New cigars table ──────────────────────────────────────────────────────
    orders = rec.get("recommended_orders", [])
    if orders:
        story.append(_section_header("NEW CIGAR RECOMMENDATIONS"))
        story.append(Spacer(1, 4))

        col_w = [0.3, 1.55, 1.2, 0.7, 0.65, 0.65, 0.5, 0.65, 0.7, 2.5]
        col_w = [w * inch for w in col_w]

        rows_new = [[
            "Rank", "Cigar", "Brand", "Vitola", "Conv.", "Branches",
            "Boxes", "Wholesale", "MSRP/st.", "Rationale",
        ]]
        for o in orders:
            conviction = o.get("conviction", "")
            cost = o.get("cost_estimate")
            msrp = o.get("msrp_per_stick")
            rows_new.append([
                str(o.get("rank", "")),
                Paragraph(o.get("name", ""), wrap),
                Paragraph(o.get("brand", ""), wrap),
                o.get("vitola", ""),
                conviction.upper(),
                Paragraph(", ".join(o.get("branches_agreed", [])), wrap),
                str(o.get("boxes", "")),
                f"${cost:,.0f}" if isinstance(cost, (int, float)) else "—",
                f"${msrp:.2f}"  if isinstance(msrp, (int, float)) else "—",
                Paragraph(o.get("rationale", ""), wrap),
            ])

        if rec.get("total_order_cost") is not None:
            total_new = rec.get("total_order_cost")
            rows_new.append([
                "", "", "", "", "", "", "TOTAL",
                f"${total_new:,.0f}" if isinstance(total_new, (int, float)) else "—",
                "", "",
            ])

        tbl_new = Table(rows_new, colWidths=col_w, repeatRows=1)
        ts_new = _col_header_style() + _row_style(len(orders))
        conv_colors = {"HIGH": GREEN, "MEDIUM": AMBER, "BOLD": RED}
        for i, o in enumerate(orders, 1):
            conv = o.get("conviction", "").upper()
            if conv in conv_colors:
                ts_new.append(("TEXTCOLOR", (4, i), (4, i), conv_colors[conv]))
                ts_new.append(("FONTNAME",  (4, i), (4, i), "Helvetica-Bold"))
        if rec.get("total_order_cost") is not None:
            last = len(rows_new) - 1
            ts_new += [
                ("BACKGROUND", (0, last), (-1, last), GOLD),
                ("FONTNAME",   (0, last), (-1, last), "Helvetica-Bold"),
            ]
        tbl_new.setStyle(TableStyle(ts_new))
        story.append(tbl_new)

        if rec.get("not_recommended"):
            story.append(Spacer(1, 4))
            passed_lines = [
                f"<b>Passed on:</b> "
                + ", ".join(
                    f"{x.get('name','?')} ({x.get('reason','')})"
                    for x in rec["not_recommended"]
                )
            ]
            story.append(Paragraph(passed_lines[0], small))

        story.append(Spacer(1, 12))

    # ── Restock table ─────────────────────────────────────────────────────────
    restock_items = restock.get("items", [])
    story.append(_section_header(
        f"LOW-STOCK REORDER  ({restock.get('flagged_count', len(restock_items))} flagged, "
        f"{horizon_days}-day window)"
    ))
    story.append(Spacer(1, 4))

    if restock_items:
        col_w_r = [0.65, 2.05, 1.25, 0.7, 0.5, 0.55, 0.6, 0.5, 0.55, 0.85]
        col_w_r = [w * inch for w in col_w_r]

        _urgency_label = {
            "critical": "🔴 Critical", "high": "🟠 High",
            "medium": "🟡 Medium",    "low":  "🟢 Low",
        }

        rows_rest = [["Urgency", "Description", "Brand", "Status",
                       "On Hand", "Days", "Vel./mo", "Boxes", "Units", "Wholesale"]]

        for item in restock_items:
            urgency = item.get("urgency", "")
            days    = item.get("days_until_stockout")
            sf      = item.get("seasonal_factor")
            sf_tag  = f" ×{sf:.2f}" if sf else ""
            cost    = item.get("reorder_cost", 0)
            rows_rest.append([
                _urgency_label.get(urgency, urgency),
                Paragraph(item.get("description", ""), wrap),
                Paragraph(item.get("brand", ""), wrap),
                item.get("status", "").replace("_", " ").capitalize(),
                str(item.get("on_hand", 0)),
                f"{int(days)}d{sf_tag}" if days is not None else f"OOS{sf_tag}",
                f"{item.get('monthly_velocity', 0):.1f}",
                str(item.get("reorder_boxes", 0)),
                str(item.get("reorder_qty", 0)),
                f"${cost:,.0f}",
            ])

        total_rest = restock.get("total_cost_ordered", 0)
        rows_rest.append(["", "", "", "", "", "", "", "TOTAL", "",
                           f"${total_rest:,.0f}"])

        tbl_rest = Table(rows_rest, colWidths=col_w_r, repeatRows=1)
        ts_rest = _col_header_style() + _row_style(len(restock_items))
        last_r = len(rows_rest) - 1
        ts_rest += [
            ("BACKGROUND", (0, last_r), (-1, last_r), GOLD),
            ("FONTNAME",   (0, last_r), (-1, last_r), "Helvetica-Bold"),
        ]
        tbl_rest.setStyle(TableStyle(ts_rest))
        story.append(tbl_rest)

        if restock.get("reasoning"):
            story.append(Spacer(1, 4))
            story.append(Paragraph(f"<b>Prioritization:</b> {restock['reasoning']}", small))

        skipped = restock.get("items_skipped", [])
        if skipped:
            story.append(Spacer(1, 6))
            story.append(Paragraph(
                f"<b>Deferred — budget exhausted ({len(skipped)} items):</b> "
                + ", ".join(item.get("description", "?") for item in skipped[:8])
                + ("…" if len(skipped) > 8 else ""),
                small,
            ))
    else:
        story.append(Paragraph("No low-stock items flagged — inventory looks healthy.", body))

    story.append(Spacer(1, 12))

    # ── By vendor table ───────────────────────────────────────────────────────
    if grouped:
        story.append(_section_header("CONSOLIDATED ORDER BY VENDOR"))
        story.append(Spacer(1, 4))

        col_w_v = [2.5, 0.5, 2.5, 0.85, 0.85]
        col_w_v = [w * inch for w in col_w_v]

        for parent, group in grouped.items():
            group_total = group.get("group_total_cost", 0)
            vendor_rows = [
                [Paragraph(f"<b>{parent}</b>", h3),
                 "", "", "",
                 Paragraph(f"<b>${group_total:,.2f}</b>", right_s)],
            ]
            restock_g = group.get("restock", [])
            new_g     = group.get("new_cigars", [])

            for item in restock_g:
                vendor_rows.append([
                    Paragraph(f"[RESTOCK] {item.get('description','')}", wrap),
                    str(item.get("reorder_boxes", 0)),
                    Paragraph(f"× {item.get('box_size',20)} sticks/box = {item.get('reorder_qty',0)} sticks", wrap),
                    item.get("urgency", "").capitalize(),
                    f"${item.get('reorder_cost',0):,.2f}",
                ])
            for order in new_g:
                vitola = order.get("vitola", "")
                vendor_rows.append([
                    Paragraph(f"[NEW] {order.get('name','')} — {vitola}", wrap),
                    str(order.get("boxes", 0)),
                    Paragraph(f"× {order.get('box_size',20)} sticks/box", wrap),
                    order.get("conviction", "").capitalize(),
                    f"${order.get('cost_estimate',0):,.2f}",
                ])

            vendor_tbl = Table(vendor_rows, colWidths=col_w_v)
            ts_v = [
                ("BACKGROUND",    (0, 0), (-1, 0), CREAM),
                ("LINEBELOW",     (0, 0), (-1, 0), 0.5, AMBER),
                ("FONTSIZE",      (0, 0), (-1, -1), 7),
                ("TOPPADDING",    (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("GRID",          (0, 0), (-1, -1), 0.3, HexColor("#DDDDDD")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, CREAM]),
            ]
            vendor_tbl.setStyle(TableStyle(ts_v))
            story.append(KeepTogether(vendor_tbl))
            story.append(Spacer(1, 4))

        grand_tbl = Table(
            [[Paragraph("<b>GRAND TOTAL</b>", bold_s), "", "", "",
              Paragraph(f"<b>${combined:,.2f}</b>" if combined else "—", right_s)]],
            colWidths=col_w_v,
        )
        grand_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), GOLD),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.3, HexColor("#CCCCCC")),
        ]))
        story.append(grand_tbl)

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 14))
    story.append(HRFlowable(width=page_width, color=HexColor("#CCCCCC"), thickness=0.5))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Generated by Smoke Shoppe AI  ·  {run_date}", small))

    doc.build(story)
    return out
